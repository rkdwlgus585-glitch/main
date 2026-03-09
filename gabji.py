# =================================================================
# 갑지(양도물건 소개) 자동 생성기 v3.0
# 프리미엄 디자인 + 건설업 양도양수 전문 기능
# =================================================================

from google import genai
from google.genai import types
import json
import os
import re
import sys
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
from utils import load_config, require_config

# =================================================================
# [설정]
# =================================================================
CONFIG = load_config({
    "BRAND_NAME": "서울건설정보",
    "MAIN_SITE": "SEOULMNA.CO.KR",
    "CONSULTANT_NAME": "강지현 실장",
    "PHONE": "010-9926-8661",
    "JSON_FILE": "service_account.json",
    "SHEET_NAME": "26양도매물",
    "GABJI_TAB": "26양도매물",
    "GABJI_LIVE_ENRICH": "0",
    "GABJI_SOURCE_URL_TEMPLATE": "http://www.nowmna.com/yangdo_view1.php?uid={uid}&page_no=1",
    "GABJI_SOURCE_TIMEOUT_SEC": "8",
})

try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass


def ensure_config(required_keys, context="gabji"):
    return require_config(CONFIG, required_keys, context=context)


def _desktop_output_dir():
    user_profile = os.environ.get("USERPROFILE", "").strip()
    home = os.path.expanduser("~")
    candidates = []
    if user_profile:
        candidates.append(os.path.join(user_profile, "Desktop"))
    candidates.append(os.path.join(home, "Desktop"))
    # Some KR Windows setups expose the folder name as "바탕 화면".
    if user_profile:
        candidates.append(os.path.join(user_profile, "바탕 화면"))
    candidates.append(os.path.join(home, "바탕 화면"))

    for path in candidates:
        if path and os.path.isdir(path):
            return path
    return user_profile or home or os.getcwd()


def _default_gabji_output_path(ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"갑지_{stamp}.{ext.lstrip('.')}"
    return os.path.join(_desktop_output_dir(), filename)


def _split_lines(value):
    text = str(value or "").strip()
    if not text:
        return []
    rows = [x.strip() for x in re.split(r"[\r\n]+", text) if x.strip()]
    return rows or [text]


def _is_truthy(value):
    return str(value or "").strip().lower() in {"1", "y", "yes", "true", "on"}


def _normalize_site_base_url(value):
    src = str(value or "").strip()
    if not src:
        return "https://seoulmna.co.kr"
    if "://" not in src:
        src = f"https://{src}"
    return src.rstrip("/")


def _dedupe_keep_order(lines):
    out = []
    seen = set()
    for line in lines:
        txt = str(line or "").strip()
        if not txt:
            continue
        key = txt.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(txt)
    return out


def _digits_only(value):
    return re.sub(r"\D+", "", str(value or ""))


def _extract_price_fragment(text):
    src = str(text or "").strip()
    if not src:
        return ""
    src = re.sub(r"^(최종|청구|양도가|매매가|입금가)\s*[:：]?\s*", "", src)

    m = re.search(r"[0-9][0-9,]*(?:\.[0-9]+)?\s*억(?:\s*[0-9][0-9,]*(?:\.[0-9]+)?\s*만)?", src)
    if m:
        return re.sub(r"\s+", "", m.group(0))

    m = re.search(r"[0-9][0-9,]*(?:\.[0-9]+)?\s*만", src)
    if m:
        return re.sub(r"\s+", "", m.group(0))

    m = re.search(r"[0-9][0-9,]*(?:\.[0-9]+)?", src)
    if m:
        plain = m.group(0).replace(",", "")
        if "억" in src:
            return f"{plain}억"
        return plain

    return ""


def extract_final_yangdo_price(raw_value):
    raw = str(raw_value or "").strip()
    if not raw or raw.lower() == "none":
        return "협의"

    parts = re.split(r"\s*(?:~|〜|∼|–|—|-|→|->|to|TO)\s*", raw)
    candidates = [x.strip() for x in parts if x.strip()] or [raw]
    for candidate in reversed(candidates):
        parsed = _extract_price_fragment(candidate)
        if parsed:
            return parsed

    return raw


def _normalize_header(value):
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _coerce_yangdo_candidate(value):
    src = str(value or "").strip()
    if not src:
        return ""

    parsed = extract_final_yangdo_price(src)
    parsed_fragment = _extract_price_fragment(parsed)
    if parsed_fragment:
        return parsed_fragment

    raw_fragment = _extract_price_fragment(src)
    if raw_fragment:
        return raw_fragment

    if "협의" in src and not re.search(r"\d", src):
        return "협의"
    return ""


def _is_price_hint_text(value):
    norm = _normalize_header(value)
    if not norm:
        return False
    hints = (
        "양도가",
        "청구양도가",
        "청구가",
        "매매가",
        "양도범위",
        "양도가범위",
        "범위값",
        "가격범위",
        "희망양도가",
        "희망가",
    )
    return any(token in norm for token in hints)


def _has_numeric_price(value):
    parsed = _coerce_yangdo_candidate(value)
    return bool(parsed and parsed != "협의" and re.search(r"\d", parsed))


def _to_eok(value):
    src = str(value or "").replace(",", "").strip()
    if not src or src in {"-", "None", "none"}:
        return None

    parsed_source = extract_final_yangdo_price(src)

    m_eok = re.search(r"([-+]?\d+(?:\.\d+)?)\s*억", parsed_source)
    if m_eok:
        result = float(m_eok.group(1))
        m_man = re.search(r"억\s*([-+]?\d+(?:\.\d+)?)\s*만", parsed_source)
        if m_man:
            result += float(m_man.group(1)) / 10000
        return result

    m_man = re.search(r"([-+]?\d+(?:\.\d+)?)\s*만", parsed_source)
    if m_man:
        return float(m_man.group(1)) / 10000

    m_num = re.search(r"[-+]?\d+(?:\.\d+)?", parsed_source)
    if not m_num:
        return None

    num = float(m_num.group(0))
    if num >= 1000 and ("억" not in parsed_source) and ("만" not in parsed_source):
        return num / 10000
    return num


def _format_eok(value):
    if value is None:
        return "-"
    txt = f"{value:.2f}".rstrip("0").rstrip(".")
    return f"{txt}억"


def _format_capital(value):
    src = str(value or "").strip()
    if not src:
        return "-"
    if "억" in src or "만" in src or "원" in src:
        return src
    if re.fullmatch(r"\d+(?:\.\d+)?", src):
        return f"{src}억"
    return src


def _format_balance(value):
    src = str(value or "").strip()
    if not src:
        return "-"
    if any(token in src for token in ["억", "만", "원"]):
        return src

    normalized = src.replace(",", "")
    if re.fullmatch(r"\d+(?:\.\d+)?", normalized):
        if "." in normalized:
            return f"{normalized}만원"
        return f"{int(float(normalized)):,}만원"
    return src


def _format_founded_year(value):
    src = str(value or "").strip()
    if not src:
        return "-"
    if "년" in src:
        return src
    if re.fullmatch(r"\d{4}", src):
        return f"{src}년"
    return src


def _parse_percent_value(text, keyword=None):
    src = str(text or "").strip()
    if not src:
        return None

    if keyword:
        pattern = rf"{re.escape(keyword)}[^0-9+-]*([-+]?\d[\d,]*(?:\.\d+)?)\s*%?"
    else:
        pattern = r"([-+]?\d[\d,]*(?:\.\d+)?)\s*%?"

    match = re.search(pattern, src)
    if not match:
        return None

    raw = match.group(1).replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return None


def _format_license_name(value):
    name = str(value or "").strip()
    if not name:
        return ""

    compact = re.sub(r"\s+", "", name)
    if compact.endswith("공사업"):
        return name
    if compact.endswith("사업"):
        return name
    if compact.endswith("공"):
        return f"{name}사업"
    return f"{name}공사업"


class ListingSheetLookup:
    COL_SEQ = 0
    COL_LICENSE = 2
    COL_LICENSE_YEAR = 3
    COL_SIPYEONG = 4
    COL_Y20 = 5
    COL_Y21 = 6
    COL_Y22 = 7
    COL_Y23 = 8
    COL_Y24 = 9
    COL_SUM3 = 10
    COL_SUM5 = 11
    COL_Y25 = 12
    COL_FOUNDED = 13
    COL_SHARES = 14
    COL_COMPANY_TYPE = 15
    COL_LOCATION = 16
    COL_BALANCE = 17
    COL_PRICE = 18
    COL_CAPITAL = 19
    COL_ASSOCIATION = 20
    COL_MEMO = 31
    COL_SOURCE_ID = 34
    PRICE_PRIMARY_HEADER_HINTS = (
        "양도가",
        "최종양도가",
        "최종가",
        "매매가",
        "양도금액",
        "양수도가",
    )
    PRICE_FALLBACK_HEADER_HINTS = (
        "청구양도가",
        "청구가",
        "청구금액",
        "양도범위",
        "양도가범위",
        "범위값",
        "가격범위",
        "희망양도가",
        "희망가",
    )
    NOTE_HEADER_HINTS = (
        "비고",
        "특이사항",
        "메모",
        "참고",
        "설명",
    )
    SOURCE_UID_HEADER_HINTS = (
        "원본uid",
        "sourceuid",
        "소스uid",
        "원본번호",
        "매물uid",
        "uid",
    )
    LIVE_MEMO_HINTS = (
        "비고",
        "특이사항",
        "주요체크사항",
    )
    SALES_YEAR_KEYS = ("2020년", "2021년", "2022년", "2023년", "2024년", "2025년")
    SENSITIVE_SKIP_TOKENS = {"-", "--", "신규"}

    def __init__(self):
        ensure_config(["JSON_FILE", "SHEET_NAME"], "gabji:sheet")
        self.json_file = str(CONFIG.get("JSON_FILE", "service_account.json")).strip()
        self.sheet_name = str(CONFIG.get("SHEET_NAME", "26양도매물")).strip()
        self.tab_name = str(CONFIG.get("GABJI_TAB", self.sheet_name)).strip() or self.sheet_name
        self._rows_cache = None
        self._header_index_cache = {}

    def _has_meaningful_text(self, value):
        text = str(value or "").replace("\xa0", " ").strip()
        if not text:
            return False
        return text.lower() not in {"none", "null"}

    def _normalize_note_line(self, line):
        text = str(line or "").replace("\xa0", " ").strip()
        if not text:
            return ""
        text = re.sub(r"(?<=\d);(?=\d)", ".", text)
        text = re.sub(r"^[\s\-\*\u2022\u25cf\u25a0\u25a1\u25c6\u25b6\u25b7\u2605\u25c7\u3000\uff0a]+", "", text)
        text = re.sub(r"\s+", " ", text).strip(" \t\r\n-:")
        return text

    def _split_html_cell_lines(self, cell):
        if cell is None:
            return []
        raw = ""
        try:
            raw = cell.decode_contents()
        except (AttributeError, TypeError):
            raw = str(cell or "")
        raw = raw.replace("&nbsp;", " ")
        raw = re.sub(r"<br\s*/?>", "\n", raw, flags=re.I)
        raw = re.sub(r"</(li|p|div|tr|td)>", "\n", raw, flags=re.I)
        text = BeautifulSoup(raw, "html.parser").get_text("\n")
        lines = [self._normalize_note_line(x) for x in re.split(r"[\r\n]+", text)]
        lines = [x for x in lines if x and x.lower() != "none"]
        return _dedupe_keep_order(lines)

    def _extract_uid_candidates(self, value):
        src = str(value or "")
        if not src:
            return []
        src = re.sub(r"<br\s*/?>", " ", src, flags=re.I)
        src = src.replace("\xa0", " ")
        candidates = []
        patterns = (
            r"uid\s*=\s*(\d{4,7})",
            r"등록번호\s*[:：]?\s*(\d{4,7})",
        )
        for pat in patterns:
            for m in re.finditer(pat, src, flags=re.I):
                candidates.append(m.group(1))
        for m in re.finditer(r"\b(\d{4,7})\b", src):
            num = m.group(1)
            if len(num) == 4:
                try:
                    year = int(num)
                    if 1900 <= year <= 2100:
                        continue
                except (ValueError, TypeError):
                    pass
            candidates.append(num)
        return _dedupe_keep_order(candidates)

    def _resolve_source_uid(self, row, registration_no=""):
        target = _digits_only(registration_no)
        candidate_cells = [
            self._row_value(row, self.COL_SOURCE_ID),
            self._row_value(row, self.COL_MEMO),
            self._row_value(row, self.COL_PRICE),
        ]
        for idx in self._header_indexes(self.SOURCE_UID_HEADER_HINTS):
            candidate_cells.append(self._row_value(row, idx))
        for idx in self._header_indexes(self.PRICE_FALLBACK_HEADER_HINTS):
            candidate_cells.append(self._row_value(row, idx))
        for idx in self._header_indexes(self.NOTE_HEADER_HINTS):
            candidate_cells.append(self._row_value(row, idx))

        uid_candidates = []
        for cell in candidate_cells:
            uid_candidates.extend(self._extract_uid_candidates(cell))
        uid_candidates = _dedupe_keep_order(uid_candidates)
        if not uid_candidates:
            return ""

        for cand in uid_candidates:
            if len(cand) >= 5 and cand != target:
                return cand
        for cand in uid_candidates:
            if cand != target:
                return cand
        return uid_candidates[0]

    def _detail_url_from_uid(self, source_uid):
        template = str(
            CONFIG.get(
                "GABJI_SOURCE_URL_TEMPLATE",
                "http://www.nowmna.com/yangdo_view1.php?uid={uid}&page_no=1",
            )
        ).strip()
        if "{uid}" in template:
            return template.format(uid=source_uid)
        sep = "&" if "?" in template else "?"
        return f"{template}{sep}uid={source_uid}"

    def _seoul_detail_url_from_wr_id(self, wr_id):
        base_url = _normalize_site_base_url(
            CONFIG.get("GABJI_SITE_URL", CONFIG.get("SITE_URL", "https://seoulmna.co.kr"))
        )
        return f"{base_url}/mna/{int(wr_id)}"

    def _fetch_seoul_industry_rows(self, wr_id):
        wr_txt = str(wr_id or "").strip()
        if not wr_txt.isdigit():
            return []

        try:
            timeout = float(str(CONFIG.get("GABJI_SOURCE_TIMEOUT_SEC", "8")).strip() or "8")
        except (ValueError, TypeError):
            timeout = 8.0

        url = self._seoul_detail_url_from_wr_id(int(wr_txt))
        try:
            resp = requests.get(
                url,
                timeout=max(2.0, timeout),
                headers={"User-Agent": "Mozilla/5.0"},
            )
            resp.raise_for_status()
            if not resp.encoding or resp.encoding.lower() in {"iso-8859-1", "ascii"}:
                resp.encoding = resp.apparent_encoding or "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")
            return self._parse_live_industry_rows(soup)
        except (requests.RequestException, ValueError, AttributeError, KeyError):
            return []

    def _find_td_value_by_labels(self, soup, labels):
        norm_labels = [_normalize_header(x).replace(":", "") for x in labels if str(x).strip()]
        if not norm_labels:
            return "", None
        candidates = []
        for td in soup.find_all("td"):
            label = str(td.get_text(" ", strip=True) or "").replace("\xa0", " ")
            norm_label = _normalize_header(label).replace(":", "")
            if not norm_label:
                continue
            matched_token = None
            match_rank = None
            for token in norm_labels:
                if not token:
                    continue
                if norm_label == token:
                    matched_token = token
                    match_rank = 0
                    break
                if norm_label.startswith(token) or norm_label.endswith(token):
                    matched_token = token
                    match_rank = 1
                    break
                if token in norm_label:
                    matched_token = token
                    match_rank = 2
                    break
            if matched_token is None:
                continue
            value_td = td.find_next_sibling("td")
            if value_td is None:
                continue
            text = str(value_td.get_text(" ", strip=True) or "").replace("\xa0", " ").strip()
            score = (match_rank, len(norm_label))
            candidates.append((score, text, value_td))
        if not candidates:
            return "", None
        candidates.sort(key=lambda x: x[0])
        for _, text, value_td in candidates:
            if self._has_meaningful_text(text):
                return text, value_td
        _, text, value_td = candidates[0]
        return text, value_td

    def _sum_sales_years(self, sales, years):
        values = []
        for year in years:
            values.append(_to_eok(str((sales or {}).get(year, "")).strip()))
        if not any(v is not None for v in values):
            return None
        return sum(v for v in values if v is not None)

    def _effective_sum_text(self, raw_sum, sales, years):
        raw_text = str(raw_sum or "").strip()
        if self._has_meaningful_text(raw_text) and raw_text not in self.SENSITIVE_SKIP_TOKENS:
            return raw_text
        numeric_sum = self._sum_sales_years(sales, years)
        return _format_eok(numeric_sum) if numeric_sum is not None else "-"

    def _sum3_year_keys(self):
        if len(self.SALES_YEAR_KEYS) >= 5:
            # seoulmna.co.kr 기준: 3년은 2022~2024 (2025 제외)
            return list(self.SALES_YEAR_KEYS[2:5])
        return list(self.SALES_YEAR_KEYS[-3:])

    def _sum5_year_keys(self):
        if len(self.SALES_YEAR_KEYS) >= 5:
            # seoulmna.co.kr 기준: 5년은 2020~2024 (2025 제외)
            return list(self.SALES_YEAR_KEYS[:5])
        return list(self.SALES_YEAR_KEYS)

    def _header_to_year_key(self, header_norm):
        value = str(header_norm or "").strip()
        if not value:
            return ""

        m4 = re.search(r"(20\d{2})", value)
        if m4:
            key = f"{m4.group(1)}년"
            return key if key in self.SALES_YEAR_KEYS else ""

        m2 = re.fullmatch(r"(20|21|22|23|24|25)", value)
        if m2:
            key = f"20{m2.group(1)}년"
            return key if key in self.SALES_YEAR_KEYS else ""

        return ""

    def _parse_live_industry_rows(self, soup):
        best_rows = []
        for table in soup.find_all("table"):
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue

            header = [
                str(c.get_text(" ", strip=True) or "").replace("\xa0", " ").strip()
                for c in rows[0].find_all(["td", "th"], recursive=False)
            ]
            if not header:
                continue

            norm_header = [_normalize_header(x) for x in header]
            if len(norm_header) < 2:
                continue

            year_index = {}
            for idx, value in enumerate(norm_header):
                year_key = self._header_to_year_key(value)
                if year_key and year_key not in year_index:
                    year_index[year_key] = idx

            def _pick_index(tokens, exclude=None):
                exclude = set(exclude or [])
                for idx, value in enumerate(norm_header):
                    if idx in exclude:
                        continue
                    if any(token in value for token in tokens):
                        return idx
                return None

            idx_license = _pick_index(["업종"])
            idx_license_year = _pick_index(["면허년도", "면허연도"], exclude={idx_license})

            idx_sum3 = None
            idx_sum5 = None
            for idx, value in enumerate(norm_header):
                if "합계" not in value:
                    continue
                if "3년" in value and idx_sum3 is None:
                    idx_sum3 = idx
                if "5년" in value and idx_sum5 is None:
                    idx_sum5 = idx

            used_for_sip = set(year_index.values())
            used_for_sip.update(x for x in [idx_license, idx_license_year, idx_sum3, idx_sum5] if isinstance(x, int))
            idx_sipyeong = _pick_index(
                ["시공능력평가액", "시공능력평가", "시공능력", "시평"],
                exclude=used_for_sip,
            )

            if idx_license is None or not year_index:
                continue

            parsed = []
            for tr in rows[1:]:
                cells = [
                    str(c.get_text(" ", strip=True) or "").replace("\xa0", " ").strip()
                    for c in tr.find_all(["td", "th"], recursive=False)
                ]
                if not cells or not any(cells):
                    continue

                license_name = cells[idx_license].strip() if idx_license < len(cells) else ""
                if not self._has_meaningful_text(license_name):
                    continue
                if any(token in license_name for token in ["재무제표", "주요체크사항", "비고", "양도리스트", "회사개요"]):
                    continue

                license_year = "-"
                if isinstance(idx_license_year, int) and idx_license_year < len(cells):
                    candidate = cells[idx_license_year].strip()
                    if self._has_meaningful_text(candidate):
                        license_year = candidate

                sales = {year: "-" for year in self.SALES_YEAR_KEYS}
                for year_key, col_idx in year_index.items():
                    if col_idx < len(cells):
                        value = cells[col_idx].strip()
                        if self._has_meaningful_text(value):
                            sales[year_key] = value

                sum3_raw = ""
                if isinstance(idx_sum3, int) and idx_sum3 < len(cells):
                    sum3_raw = cells[idx_sum3].strip()
                sum5_raw = ""
                if isinstance(idx_sum5, int) and idx_sum5 < len(cells):
                    sum5_raw = cells[idx_sum5].strip()

                # Live 테이블이 우측 컬럼을 누락하는 경우(예: 2025년 값이 뒤에서 2번째로 밀림) 보정
                if sales.get("2025년", "-") in self.SENSITIVE_SKIP_TOKENS and len(cells) >= 2:
                    tail_candidate = str(cells[-2] or "").strip()
                    if self._has_meaningful_text(tail_candidate) and tail_candidate not in {"-", "--"}:
                        if not re.search(r"\d", tail_candidate) or "신규" in tail_candidate:
                            sales["2025년"] = tail_candidate
                            if str(sum5_raw or "").strip() == tail_candidate:
                                sum5_raw = ""

                sipyeong = "-"
                if isinstance(idx_sipyeong, int) and idx_sipyeong < len(cells):
                    candidate = cells[idx_sipyeong].strip()
                    if self._has_meaningful_text(candidate):
                        sipyeong = candidate

                parsed.append(
                    {
                        "업종": license_name,
                        "면허년도": license_year,
                        "시공능력평가액": sipyeong,
                        "매출": sales,
                        "3년합계": self._effective_sum_text(sum3_raw, sales, self._sum3_year_keys()),
                        "5년합계": self._effective_sum_text(sum5_raw, sales, self._sum5_year_keys()),
                    }
                )

            if len(parsed) > len(best_rows):
                best_rows = parsed

        return best_rows

    def _parse_live_listing_detail(self, html_text):
        soup = BeautifulSoup(str(html_text or ""), "html.parser")
        page_text = soup.get_text(" ", strip=True).replace("\xa0", " ")
        detail = {}

        m_reg = re.search(r"등록번호\s*[:：]?\s*(\d{4,7})", page_text)
        if m_reg:
            detail["원본등록번호"] = m_reg.group(1)

        company_type, _ = self._find_td_value_by_labels(soup, ["회사형태", "회사"])
        capital, _ = self._find_td_value_by_labels(soup, ["자본금"])
        location, _ = self._find_td_value_by_labels(soup, ["소재지"])
        founded, _ = self._find_td_value_by_labels(soup, ["법인설립일", "법인년도"])
        association, _ = self._find_td_value_by_labels(soup, ["협회가입", "협회"])
        price, _ = self._find_td_value_by_labels(soup, ["양도가", "매매가", "최종 양도가", "최종가"])
        shares, _ = self._find_td_value_by_labels(soup, ["공제조합출자좌수", "출자좌수", "출자수", "좌수"])
        balance, _ = self._find_td_value_by_labels(soup, ["대출후남은잔액", "대출후잔액", "공제잔액", "잔액"])
        memo_text, memo_cell = self._find_td_value_by_labels(soup, self.LIVE_MEMO_HINTS)

        if shares and "/" in shares:
            parts = [x.strip() for x in shares.split("/", 1)]
            shares = parts[0] if parts else shares
            if len(parts) > 1 and not self._has_meaningful_text(balance):
                balance = parts[1]

        if self._has_meaningful_text(company_type):
            detail["회사형태"] = company_type
        if self._has_meaningful_text(capital):
            detail["자본금"] = _format_capital(capital.replace(" ", ""))
        if self._has_meaningful_text(location):
            detail["소재지"] = location
        if self._has_meaningful_text(founded):
            detail["법인설립일"] = _format_founded_year(founded.replace(" ", ""))
        if self._has_meaningful_text(association):
            detail["협회가입"] = association
        if self._has_meaningful_text(price):
            detail["양도가"] = extract_final_yangdo_price(price)
        if self._has_meaningful_text(shares):
            share_text = str(shares).strip()
            if re.search(r"\d", share_text):
                if "좌" not in share_text and share_text not in {"-", "--"}:
                    share_text = f"{share_text}좌"
                detail["공제조합출자좌수"] = share_text
        if self._has_meaningful_text(balance):
            balance_text = _format_balance(balance.replace(" ", ""))
            if re.search(r"\d", balance_text):
                detail["공제조합잔액"] = balance_text

        memo_lines = []
        if memo_cell is not None:
            memo_lines = self._split_html_cell_lines(memo_cell)
        elif self._has_meaningful_text(memo_text):
            memo_lines = _dedupe_keep_order(
                [
                    self._normalize_note_line(x)
                    for x in _split_lines(memo_text)
                    if self._normalize_note_line(x)
                ]
            )
        if memo_lines:
            detail["비고"] = memo_lines

        industry_rows = self._parse_live_industry_rows(soup)
        if industry_rows:
            detail["업종정보"] = industry_rows

        return detail

    def _fetch_live_listing_detail(self, source_uid):
        if not source_uid:
            return {}
        try:
            timeout = float(str(CONFIG.get("GABJI_SOURCE_TIMEOUT_SEC", "8")).strip() or "8")
        except (ValueError, TypeError):
            timeout = 8.0

        url = self._detail_url_from_uid(source_uid)
        try:
            resp = requests.get(
                url,
                timeout=max(2.0, timeout),
                headers={"User-Agent": "Mozilla/5.0"},
            )
            resp.raise_for_status()
            content_head = resp.content[:2048].lower()
            if b"charset=euc-kr" in content_head:
                resp.encoding = "euc-kr"
            elif not resp.encoding or resp.encoding.lower() in {"iso-8859-1", "ascii"}:
                resp.encoding = resp.apparent_encoding or "utf-8"

            detail = self._parse_live_listing_detail(resp.text)
            if not detail:
                return {}
            detail["원본UID"] = str(source_uid)
            detail["원본URL"] = url
            return detail
        except (requests.RequestException, ValueError, AttributeError, KeyError):
            return {}

    def _industry_key(self, row):
        name = str((row or {}).get("업종", "")).strip()
        if not name:
            return ""
        return re.sub(r"\s+", "", name.lower())

    def _normalize_industry_row(self, row):
        src = row if isinstance(row, dict) else {}
        sales_src = src.get("매출", {}) if isinstance(src.get("매출", {}), dict) else {}
        sales = {}
        for y in self.SALES_YEAR_KEYS:
            v = str(sales_src.get(y, "")).strip()
            sales[y] = v if self._has_meaningful_text(v) else "-"
        return {
            "업종": str(src.get("업종", "")).strip(),
            "면허년도": str(src.get("면허년도", "")).strip() or "-",
            "시공능력평가액": str(src.get("시공능력평가액", "")).strip() or "-",
            "매출": sales,
            "3년합계": str(src.get("3년합계", "")).strip() or "-",
            "5년합계": str(src.get("5년합계", "")).strip() or "-",
        }

    def _normalized_digits(self, value):
        return re.sub(r"\D+", "", str(value or ""))

    def _can_apply_sensitive_value(self, current, incoming):
        incoming_txt = str(incoming or "").strip()
        if not self._has_meaningful_text(incoming_txt) or incoming_txt in self.SENSITIVE_SKIP_TOKENS:
            return False

        current_txt = str(current or "").strip()
        if not self._has_meaningful_text(current_txt) or current_txt in self.SENSITIVE_SKIP_TOKENS:
            return True

        cur_digits = self._normalized_digits(current_txt)
        in_digits = self._normalized_digits(incoming_txt)
        if cur_digits and in_digits and cur_digits == in_digits:
            return True
        return False

    def _is_verified_sum(self, sales, raw_sum, years):
        txt = str(raw_sum or "").strip()
        if not self._has_meaningful_text(txt) or txt in self.SENSITIVE_SKIP_TOKENS:
            return False
        parsed = _to_eok(txt)
        if parsed is None:
            return False
        calc = self._sum_sales_years(sales, years)
        if calc is None:
            return False
        return abs(parsed - calc) <= 0.11

    def _merge_industry_rows(self, base_rows, live_rows):
        base_rows = base_rows if isinstance(base_rows, list) else []
        live_rows = live_rows if isinstance(live_rows, list) else []
        merged = {}
        order = []

        # 업종 출력 순서는 홈페이지(라이브) 순서를 우선한다.
        for row in live_rows:
            norm = self._normalize_industry_row(row)
            key = self._industry_key(norm)
            if not key:
                continue
            if key not in merged:
                merged[key] = norm
                order.append(key)

        for row in base_rows:
            norm = self._normalize_industry_row(row)
            key = self._industry_key(norm)
            if not key:
                continue
            if key not in merged:
                merged[key] = norm
                order.append(key)
                continue

            cur = merged[key]
            for field in ["면허년도", "시공능력평가액"]:
                incoming = str(norm.get(field, "")).strip()
                current = str(cur.get(field, "")).strip()
                if self._has_meaningful_text(incoming) and incoming not in {"-", "--"}:
                    if not self._has_meaningful_text(current) or current in {"-", "--", "신규"}:
                        cur[field] = incoming

            cur_sales = cur.get("매출", {}) if isinstance(cur.get("매출"), dict) else {}
            in_sales = norm.get("매출", {}) if isinstance(norm.get("매출"), dict) else {}
            for y in self.SALES_YEAR_KEYS:
                incoming = str(in_sales.get(y, "")).strip()
                current = str(cur_sales.get(y, "")).strip()
                if self._can_apply_sensitive_value(current, incoming):
                    cur_sales[y] = incoming
                elif not self._has_meaningful_text(current):
                    cur_sales[y] = "-"
            cur["매출"] = cur_sales

            cur_sum3 = str(cur.get("3년합계", "")).strip()
            in_sum3 = str(norm.get("3년합계", "")).strip()
            if (not self._has_meaningful_text(cur_sum3) or cur_sum3 in self.SENSITIVE_SKIP_TOKENS) and self._is_verified_sum(
                cur_sales,
                in_sum3,
                self._sum3_year_keys(),
            ):
                cur["3년합계"] = in_sum3

            cur_sum5 = str(cur.get("5년합계", "")).strip()
            in_sum5 = str(norm.get("5년합계", "")).strip()
            if (not self._has_meaningful_text(cur_sum5) or cur_sum5 in self.SENSITIVE_SKIP_TOKENS) and self._is_verified_sum(
                cur_sales,
                in_sum5,
                self._sum5_year_keys(),
            ):
                cur["5년합계"] = in_sum5

        out = []
        for key in order:
            row = merged.get(key)
            if not row:
                continue
            if not self._has_meaningful_text(row.get("업종")):
                continue
            out.append(row)
        return out

    def _merge_live_detail(self, base_data, live_detail):
        merged = dict(base_data or {})
        if not live_detail:
            return merged

        overwrite_keys = (
            "회사형태",
            "자본금",
            "소재지",
            "법인설립일",
            "협회가입",
            "공제조합출자좌수",
            "원본UID",
            "원본URL",
            "원본등록번호",
        )
        for key in overwrite_keys:
            val = live_detail.get(key)
            if not self._has_meaningful_text(val):
                continue
            txt = str(val).strip()
            if txt in {"-", "--"}:
                continue
            merged[key] = txt

        live_balance = str(live_detail.get("공제조합잔액", "")).strip()
        cur_balance = str(merged.get("공제조합잔액", "")).strip()
        if self._can_apply_sensitive_value(cur_balance, live_balance):
            merged["공제조합잔액"] = live_balance

        # 양도가는 실무상 숫자값이 더 유효하므로:
        # - 원본이 숫자면 원본값 채택
        # - 원본이 협의이고 시트가 숫자면 시트값 유지
        # - 둘 다 숫자가 아니면 원본 협의 사용
        base_price = str(merged.get("양도가", "")).strip()
        live_price = str(live_detail.get("양도가", "")).strip()
        if self._has_meaningful_text(live_price):
            live_price = extract_final_yangdo_price(live_price)
            if _has_numeric_price(live_price):
                merged["양도가"] = live_price
            elif "협의" in live_price and not _has_numeric_price(base_price):
                merged["양도가"] = "협의"
            elif not _has_numeric_price(base_price):
                merged["양도가"] = live_price

        base_industry = merged.get("업종정보", [])
        live_industry = live_detail.get("업종정보", [])
        merged_industry = self._merge_industry_rows(base_industry, live_industry)
        if merged_industry:
            merged["업종정보"] = merged_industry

        live_notes = live_detail.get("비고", [])
        if isinstance(live_notes, str):
            live_notes = _split_lines(live_notes)
        live_notes = [self._normalize_note_line(x) for x in live_notes if self._normalize_note_line(x)]

        base_notes = merged.get("비고", [])
        if isinstance(base_notes, str):
            base_notes = _split_lines(base_notes)
        base_notes = [self._normalize_note_line(x) for x in base_notes if self._normalize_note_line(x)]

        merged_notes = _dedupe_keep_order(live_notes + base_notes)
        if merged_notes:
            merged["비고"] = merged_notes
            admin_keys = ("행정", "처분", "결손", "기술자", "출자", "면허", "신용", "부채", "유동")
            admin_lines = [x for x in merged_notes if any(k in x for k in admin_keys)]
            if admin_lines:
                merged["행정사항"] = admin_lines

        return merged

    def _row_value(self, row, idx):
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx]).strip()

    def _pick(self, values, idx):
        if idx < len(values):
            return values[idx]
        if len(values) == 1:
            return values[0]
        return ""

    def _header_indexes(self, hints):
        key = tuple(hints)
        cached = self._header_index_cache.get(key)
        if cached is not None:
            return list(cached)

        rows = self._load_rows()
        if not rows:
            self._header_index_cache[key] = []
            return []

        headers = rows[0]
        norm_hints = [_normalize_header(x) for x in hints if _normalize_header(x)]
        matched = []
        for idx, header in enumerate(headers):
            norm_header = _normalize_header(header)
            if not norm_header:
                continue
            if any(hint in norm_header for hint in norm_hints):
                matched.append(idx)

        self._header_index_cache[key] = matched
        return list(matched)

    def _resolve_yangdo_price(self, row):
        candidates = [
            self.COL_PRICE,
            *self._header_indexes(self.PRICE_PRIMARY_HEADER_HINTS),
            *self._header_indexes(self.PRICE_FALLBACK_HEADER_HINTS),
            self.COL_MEMO,
            *self._header_indexes(self.NOTE_HEADER_HINTS),
        ]

        seen = set()
        ordered_indexes = []
        for idx in candidates:
            if not isinstance(idx, int) or idx < 0:
                continue
            if idx in seen:
                continue
            seen.add(idx)
            ordered_indexes.append(idx)

        consult_seen = False
        for idx in ordered_indexes:
            parsed = _coerce_yangdo_candidate(self._row_value(row, idx))
            if not parsed:
                continue
            if parsed == "협의":
                consult_seen = True
                continue
            return parsed

        # Fallback: header 위치가 엇갈린 경우, 행 텍스트 중 '청구 양도가/범위값' 힌트 셀에서 마지막 가격 강제 추출
        for cell in row:
            if not _is_price_hint_text(cell):
                continue
            parsed = _coerce_yangdo_candidate(cell)
            if not parsed:
                continue
            if parsed == "협의":
                consult_seen = True
                continue
            return parsed

        if consult_seen:
            return "협의"
        return "협의"

    def _load_rows(self):
        if self._rows_cache is not None:
            return self._rows_cache

        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(self.json_file, scope)
        client = gspread.authorize(creds)
        sheet = client.open(self.sheet_name)
        ws = sheet.worksheet(self.tab_name)
        self._rows_cache = ws.get_all_values()
        return self._rows_cache

    def _find_row(self, registration_no):
        target_raw = str(registration_no or "").strip()
        if not target_raw:
            return None, None
        target_digits = _digits_only(target_raw)

        rows = self._load_rows()
        for row_idx, row in enumerate(rows[1:], start=2):
            seq = self._row_value(row, self.COL_SEQ)
            source_id = self._row_value(row, self.COL_SOURCE_ID)

            if target_raw == seq:
                return row_idx, row
            if target_raw == source_id:
                return row_idx, row
            if target_digits and target_digits == _digits_only(seq):
                return row_idx, row
            if target_digits and target_digits == _digits_only(source_id):
                return row_idx, row

        # 정확 매칭이 없으면 소스 URL/텍스트 포함 매칭으로 2차 시도
        if target_digits:
            for row_idx, row in enumerate(rows[1:], start=2):
                row_blob = " ".join(row[: self.COL_SOURCE_ID + 1])
                if target_digits in _digits_only(row_blob):
                    return row_idx, row

        return None, None

    def load_listing(self, registration_no):
        row_idx, row = self._find_row(registration_no)
        if not row:
            raise ValueError(f"등록번호 '{registration_no}'에 해당하는 매물을 찾지 못했습니다.")
        return self._row_to_gabji_data(registration_no, row_idx, row)

    def _row_to_gabji_data(self, registration_no, row_idx, row):
        licenses = _split_lines(self._row_value(row, self.COL_LICENSE))
        years = _split_lines(self._row_value(row, self.COL_LICENSE_YEAR))
        sipyeong = _split_lines(self._row_value(row, self.COL_SIPYEONG))

        y20 = _split_lines(self._row_value(row, self.COL_Y20))
        y21 = _split_lines(self._row_value(row, self.COL_Y21))
        y22 = _split_lines(self._row_value(row, self.COL_Y22))
        y23 = _split_lines(self._row_value(row, self.COL_Y23))
        y24 = _split_lines(self._row_value(row, self.COL_Y24))
        sum3_list = _split_lines(self._row_value(row, self.COL_SUM3))
        sum5_list = _split_lines(self._row_value(row, self.COL_SUM5))
        y25 = _split_lines(self._row_value(row, self.COL_Y25))

        count = max(len(licenses), len(years), len(sipyeong), 1)
        industry_rows = []
        for idx in range(count):
            lic = self._pick(licenses, idx) or "미확인"
            year = self._pick(years, idx) or "-"
            sp = self._pick(sipyeong, idx) or "-"

            sales_2020 = self._pick(y20, idx) or "-"
            sales_2021 = self._pick(y21, idx) or "-"
            sales_2022 = self._pick(y22, idx) or "-"
            sales_2023 = self._pick(y23, idx) or "-"
            sales_2024 = self._pick(y24, idx) or "-"
            sales_2025 = self._pick(y25, idx) or "-"
            sales = {
                "2020년": sales_2020,
                "2021년": sales_2021,
                "2022년": sales_2022,
                "2023년": sales_2023,
                "2024년": sales_2024,
                "2025년": sales_2025,
            }

            raw_sum3 = self._pick(sum3_list, idx) or "-"
            raw_sum5 = self._pick(sum5_list, idx) or "-"
            sum3_text = self._effective_sum_text(raw_sum3, sales, self._sum3_year_keys())
            sum5_text = self._effective_sum_text(raw_sum5, sales, self._sum5_year_keys())

            industry_rows.append(
                {
                    "업종": lic,
                    "면허년도": year,
                    "시공능력평가액": sp,
                    "매출": sales,
                    "3년합계": sum3_text,
                    "5년합계": sum5_text,
                }
            )

        final_price = self._resolve_yangdo_price(row)

        memo = self._row_value(row, self.COL_MEMO)
        memo_lines = [x for x in _split_lines(memo) if x and x.lower() != "none"]
        admin_lines = [x for x in memo_lines if any(k in x for k in ["행정", "처분", "결손", "기술자", "출자", "면허"])]

        shares = self._row_value(row, self.COL_SHARES)
        if shares and "좌" not in shares:
            shares = f"{shares}좌"

        data = {
            "업종정보": industry_rows,
            "회사형태": self._row_value(row, self.COL_COMPANY_TYPE) or "-",
            "자본금": _format_capital(self._row_value(row, self.COL_CAPITAL)),
            "소재지": self._row_value(row, self.COL_LOCATION) or "-",
            "법인설립일": _format_founded_year(self._row_value(row, self.COL_FOUNDED)),
            "등록번호": str(registration_no or self._row_value(row, self.COL_SEQ)),
            "협회가입": self._row_value(row, self.COL_ASSOCIATION) or "-",
            "공제조합출자좌수": shares or "-",
            "공제조합잔액": _format_balance(self._row_value(row, self.COL_BALANCE)),
            "양도가": final_price,
            "행정사항": admin_lines,
            "비고": memo_lines,
            "시트행": row_idx,
        }

        if _is_truthy(CONFIG.get("GABJI_LIVE_ENRICH", "0")):
            source_uid = self._resolve_source_uid(row, registration_no=registration_no)
            if source_uid:
                live_detail = self._fetch_live_listing_detail(source_uid)
                if live_detail:
                    data = self._merge_live_detail(data, live_detail)

        # 실적 출력 순서는 서울건설정보 홈페이지(/mna/{wr_id})를 우선한다.
        site_wr_id = _digits_only(self._row_value(row, self.COL_SEQ) or registration_no)
        if site_wr_id:
            site_rows = self._fetch_seoul_industry_rows(site_wr_id)
            if site_rows:
                # 갑지는 서울건설정보 홈페이지 표시 순서를 단일 기준으로 사용한다.
                data["업종정보"] = [self._normalize_industry_row(x) for x in site_rows if isinstance(x, dict)]

        return data


# =================================================================
# [Core] 이미지 분석 엔진
# =================================================================
class GabjiGenerator:
    def __init__(self):
        ensure_config(["GEMINI_API_KEY"], "gabji:generator")
        self.client = genai.Client(api_key=CONFIG["GEMINI_API_KEY"])
    
    def analyze_image(self, image_path: str) -> dict:
        with open(image_path, "rb") as f:
            image_data = f.read()
        
        ext = os.path.splitext(image_path)[1].lower()
        mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.webp': 'image/webp'}
        mime_type = mime_map.get(ext, 'image/png')
        
        extraction_prompt = """이 이미지는 건설업체 정보 화면입니다. 
모든 정보를 정확하게 추출하여 JSON으로 반환하세요.

추출 항목:
1. 업종정보: 업종명, 면허년도, 시공능력평가액(시평), 연도별 매출(2020~2025), 3년합계, 5년합계
2. 회사정보: 회사형태, 자본금, 소재지, 법인설립일, 등록번호
3. 협회: 가입여부 (미가입/일반협회/전문협회/둘다)
4. 공제조합: 출자좌수, 잔액
5. 양도가
6. 행정사항: 특이사항 목록
7. 비고/체크사항: 부채비율, 유동비율, 행정처분, 기술자, 결손금 등 모든 정보

JSON 형식:
{
    "업종정보": [{"업종": "건축", "면허년도": "2022", "시공능력평가액": "17", 
        "매출": {"2020년": "2.1", "2021년": "-", "2022년": "3.5", "2023년": "4.1", "2024년": "1.5", "2025년": "0.9"}, 
        "3년합계": "9.1억", "5년합계": "11.2억"}],
    "회사형태": "주식회사",
    "자본금": "4.6억",
    "소재지": "지방",
    "법인설립일": "2007년",
    "등록번호": "12345",
    "협회가입": "미가입",
    "공제조합출자좌수": "107좌",
    "공제조합잔액": "5,500만원",
    "양도가": "",
    "행정사항": ["2026년 신규 면허", "공제조합 출자 5,000만원 별도"],
    "비고": ["9월 결산법인", "자체결산", "행정처분이력 無", "부채:17% 유동:619%", "기술자2명", "외부신용등급BB-"]
}"""

        try:
            print("🔍 이미지 분석 중...")
            response = self.client.models.generate_content(
                model='gemini-2.0-flash',
                contents=[types.Part.from_bytes(data=image_data, mime_type=mime_type), extraction_prompt],
                config=types.GenerateContentConfig(response_mime_type="application/json")
            )
            
            raw_text = response.text.strip()
            if raw_text.startswith("```json"): raw_text = raw_text[7:]
            if raw_text.startswith("```"): raw_text = raw_text[3:]
            if raw_text.endswith("```"): raw_text = raw_text[:-3]
            
            data = json.loads(raw_text.strip())
            print("✅ 정보 추출 완료!")
            return data
        except Exception as e:
            print(f"❌ 분석 오류: {e}")
            return None


# =================================================================
# [분석] 매물 평가 분석
# =================================================================
def analyze_property(data: dict) -> dict:
    """매물의 장단점 및 리스크 분석"""
    pros = []  # 장점
    cons = []  # 단점/주의사항
    
    비고 = data.get("비고", [])
    if isinstance(비고, str): 비고 = [비고]
    비고_text = " ".join(비고).lower()
    
    # 장점 분석
    if "행정처분" in 비고_text and "무" in 비고_text:
        pros.append("행정처분 이력 없음")
    
    # 유동비율 분석
    for item in 비고:
        if "유동" in item:
            try:
                ratio = _parse_percent_value(item, "유동")
                if ratio is not None:
                    if ratio >= 100:
                        pros.append(f"유동비율 양호 ({ratio}%)")
                    else:
                        cons.append(f"유동비율 주의 ({ratio}%)")
            except (ValueError, TypeError):
                pass

    # 부채비율 분석
    for item in 비고:
        if "부채" in item:
            try:
                ratio = _parse_percent_value(item, "부채")
                if ratio is not None:
                    if ratio <= 100:
                        pros.append(f"부채비율 양호 ({ratio}%)")
                    elif ratio <= 200:
                        cons.append(f"부채비율 보통 ({ratio}%)")
                    else:
                        cons.append(f"부채비율 높음 ({ratio}%)")
            except (ValueError, TypeError):
                pass
    
    # 기술자 분석
    if "기술자" in 비고_text:
        if "없음" in 비고_text or "이상없음" in 비고_text:
            pros.append("기술자 요건 충족")
        elif "사업자있음" in 비고_text:
            cons.append("기술자 겸업 있음")
    
    # 결손금 분석 - 실질자본금 미달인 경우만 리스크
    if "실질자본" in 비고_text and "미달" in 비고_text:
        cons.append("실질자본금 미달")
    
    # 신용등급
    if "신용등급" in 비고_text or "bb" in 비고_text.lower():
        pros.append("외부신용등급 보유")
    
    # 매출 분석
    for info in data.get("업종정보", []):
        합계3 = info.get("3년합계", "")
        if 합계3:
            try:
                val = float(''.join(filter(lambda x: x.isdigit() or x=='.', str(합계3))))
                if val >= 10:
                    pros.append("3년 매출실적 우수")
            except (ValueError, TypeError):
                pass
    
    return {"장점": pros[:4], "단점": cons[:4]}


# =================================================================
# [Output] PNG 이미지 - 프리미엄 디자인 v3
# =================================================================
class PNGGenerator:
    def generate(self, data: dict, yangdo_price: str = None, output_path: str = None) -> str:
        today = datetime.now().strftime("%Y. %m. %d.")
        brand = CONFIG["BRAND_NAME"]
        site = CONFIG["MAIN_SITE"]
        consultant = CONFIG["CONSULTANT_NAME"]
        phone = CONFIG["PHONE"]
        
        # None/빈값을 "-"로 변환하는 헬퍼 함수
        def clean(val, default="-"):
            if val is None or str(val).strip() == "" or str(val).lower() == "none":
                return default
            return str(val)
        
        price = yangdo_price if yangdo_price else data.get("양도가", "협의")
        if not price or str(price).strip() == "" or str(price).lower() == "none": 
            price = "협의"
        
        # 매물 분석
        analysis = analyze_property(data)
        
        # 장점 HTML
        pros_html = ""
        for p in analysis["장점"]:
            pros_html += f'<span class="tag tag-good">✓ {p}</span>'
        
        # 주의사항 HTML
        cons_html = ""
        for c in analysis["단점"]:
            cons_html += f'<span class="tag tag-warn">! {c}</span>'
        
        # 업종 목록
        업종_html = ""
        for info in data.get("업종정보", []):
            시평값 = clean(info.get("시공능력평가액") or info.get("시평"))
            업종_html += f'''<div class="license-card">
                <div class="license-type">{clean(_format_license_name(info.get("업종")))}</div>
                <div class="license-meta">
                    <span>면허 {clean(info.get("면허년도"))}년</span>
                    <span class="divider">|</span>
                    <span>시평 <b>{시평값}</b></span>
                </div>
            </div>'''
        
        # 매출 테이블
        table_rows = ""
        for info in data.get("업종정보", []):
            매출 = info.get("매출", {})
            시평 = clean(info.get("시공능력평가액") or info.get("시평"))
            table_rows += f'''<tr>
                <td class="cell-label">{clean(info.get("업종"))}</td>
                <td>{clean(info.get("면허년도"))}</td>
                <td class="cell-blue">{시평}</td>
                <td>{clean(매출.get("2020년"))}</td>
                <td>{clean(매출.get("2021년"))}</td>
                <td>{clean(매출.get("2022년"))}</td>
                <td>{clean(매출.get("2023년"))}</td>
                <td>{clean(매출.get("2024년"))}</td>
                <td>{clean(매출.get("2025년"))}</td>
                <td class="cell-pink">{clean(info.get("3년합계"))}</td>
                <td class="cell-pink">{clean(info.get("5년합계"))}</td>
            </tr>'''
        
        # 행정사항
        행정 = data.get("행정사항", [])
        if isinstance(행정, str): 행정 = [행정]
        행정_html = "".join([f'<li>{item}</li>' for item in 행정 if item and str(item).lower() != "none"])
        
        # 비고
        비고 = data.get("비고", [])
        if isinstance(비고, str): 비고 = [비고]
        비고_html = "".join([f'<span class="check-item">✓ {item}</span>' for item in 비고 if item and str(item).lower() != "none"])
        
        # 공제조합 정보 - "53좌/2000만" 형식 처리
        출자좌수_raw = data.get("공제조합출자좌수", "")
        잔액_raw = data.get("공제조합잔액", "")
        
        # "53좌/2000만" 형식인 경우 분리
        if 출자좌수_raw and "/" in str(출자좌수_raw):
            parts = str(출자좌수_raw).split("/")
            출자좌수 = clean(parts[0].strip())
            if len(parts) > 1 and (not 잔액_raw or str(잔액_raw) == "-"):
                잔액 = clean(parts[1].strip())
            else:
                잔액 = clean(잔액_raw)
        else:
            출자좌수 = clean(출자좌수_raw)
            잔액 = clean(잔액_raw)
        
        # 잔액 단위 정리
        if 잔액 != "-" and "만" not in 잔액 and "원" not in 잔액:
            잔액 = f"{잔액}만원"
        
        html = f'''<!DOCTYPE html>
<html>
<head>

    <meta charset="UTF-8">
    <style>
        @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Pretendard', -apple-system, sans-serif; background: #f0f2f5; }}
        
        #container {{
            width: 800px;
            background: #fff;
        }}
        
        /* 헤더 */
        .header {{
            background: linear-gradient(135deg, #1e3a5f 0%, #0d2137 100%);
            padding: 32px 40px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .header-left {{}}
        .header-brand {{
            font-size: 11px;
            color: #c9a962;
            letter-spacing: 3px;
            margin-bottom: 8px;
            font-weight: 500;
        }}
        .header-title {{
            font-size: 22px;
            color: #fff;
            font-weight: 600;
        }}
        .header-right {{
            text-align: right;
        }}
        .header-date {{
            font-size: 13px;
            color: rgba(255,255,255,0.6);
        }}
        
        /* 업종 카드 */
        .license-section {{
            padding: 24px 40px;
            background: linear-gradient(180deg, #f8f9fb 0%, #fff 100%);
            display: flex;
            gap: 16px;
            flex-wrap: wrap;
        }}
        .license-card {{
            background: linear-gradient(135deg, #1e3a5f 0%, #2d4a6f 100%);
            padding: 16px 24px;
            border-radius: 8px;
            min-width: 200px;
        }}
        .license-type {{
            font-size: 18px;
            font-weight: 700;
            color: #fff;
            margin-bottom: 6px;
        }}
        .license-meta {{
            font-size: 12px;
            color: rgba(255,255,255,0.7);
        }}
        .license-meta b {{ color: #c9a962; }}
        .divider {{ margin: 0 8px; opacity: 0.4; }}
        
        /* 분석 태그 */
        .analysis-section {{
            padding: 20px 40px;
            background: #fff;
            border-bottom: 1px solid #eee;
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
        }}
        .tag {{
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 500;
        }}
        .tag-good {{
            background: #e8f5e9;
            color: #2e7d32;
            border: 1px solid #a5d6a7;
        }}
        .tag-warn {{
            background: #fff3e0;
            color: #e65100;
            border: 1px solid #ffcc80;
        }}
        
        /* 기본 정보 그리드 */
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            padding: 24px 40px;
            gap: 0;
            border-bottom: 1px solid #eee;
        }}
        .info-item {{
            padding: 16px 0;
            border-bottom: 1px solid #f5f5f5;
        }}
        .info-item:nth-child(4n-2), .info-item:nth-child(4n-1), .info-item:nth-child(4n) {{
            border-left: 1px solid #f5f5f5;
            padding-left: 20px;
        }}
        .info-label {{
            font-size: 11px;
            color: #888;
            margin-bottom: 4px;
            font-weight: 500;
        }}
        .info-value {{
            font-size: 15px;
            color: #1a1a1a;
            font-weight: 600;
        }}
        .info-value.highlight {{
            color: #d32f2f;
        }}
        
        /* 매출 테이블 */
        .table-section {{
            padding: 24px 40px;
            border-bottom: 1px solid #eee;
        }}
        .table-title {{
            font-size: 13px;
            font-weight: 600;
            color: #1e3a5f;
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .table-title::before {{
            content: '';
            width: 4px; height: 14px;
            background: #c9a962;
            display: inline-block;
        }}
        .table-unit {{
            margin-left: auto;
            font-size: 11px;
            color: #888;
            font-weight: 400;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        th {{
            background: #1e3a5f;
            color: #fff;
            padding: 12px 6px;
            font-weight: 500;
            text-align: center;
        }}
        td {{
            padding: 12px 6px;
            text-align: center;
            border-bottom: 1px solid #eee;
            color: #444;
            white-space: nowrap;
        }}
        .cell-label {{ font-weight: 600; color: #1e3a5f; text-align: left; padding-left: 12px; white-space: nowrap; }}
        .cell-blue {{ color: #1565c0; font-weight: 600; }}
        .cell-pink {{ color: #d32f2f; font-weight: 700; background: #fef0f0; }}
        
        /* 행정사항 */
        .admin-section {{
            padding: 20px 40px;
            background: #fffde7;
            border-bottom: 1px solid #fff59d;
        }}
        .section-header {{
            font-size: 12px;
            font-weight: 600;
            color: #f57f17;
            margin-bottom: 12px;
        }}
        .admin-section ul {{
            list-style: none;
            padding: 0;
            margin: 0;
        }}
        .admin-section li {{
            font-size: 13px;
            color: #5d4037;
            padding: 6px 0;
            padding-left: 16px;
            position: relative;
        }}
        .admin-section li::before {{
            content: '▸';
            position: absolute;
            left: 0;
            color: #ff8f00;
        }}
        
        /* 체크사항 */
        .check-section {{
            padding: 20px 40px;
            background: #e3f2fd;
            border-bottom: 1px solid #90caf9;
        }}
        .check-header {{
            font-size: 12px;
            font-weight: 600;
            color: #1565c0;
            margin-bottom: 12px;
        }}
        .check-items {{
            display: flex;
            flex-wrap: wrap;
            gap: 8px 16px;
        }}
        .check-item {{
            font-size: 12px;
            color: #1565c0;
        }}
        
        /* 양도가 */
        .price-section {{
            background: linear-gradient(135deg, #1e3a5f 0%, #0d2137 100%);
            padding: 40px;
            text-align: center;
            position: relative;
        }}
        .price-label {{
            font-size: 24px;
            color: #c9a962;
            letter-spacing: 6px;
            margin-bottom: 16px;
            font-weight: 500;
        }}
        .price-value {{
            font-size: 38px;
            color: #fff;
            font-weight: 600;
            letter-spacing: 1px;
        }}
        .price-won {{
            font-size: 26px;
            color: #c9a962;
            margin-right: 6px;
        }}
        
        /* 푸터 */
        .footer {{
            padding: 24px 40px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: #fafafa;
            border-top: 1px solid #eee;
        }}
        .footer-brand {{
            font-size: 16px;
            font-weight: 700;
            color: #1e3a5f;
        }}
        .footer-site {{
            font-size: 12px;
            color: #888;
            margin-top: 4px;
        }}
        .footer-contact {{
            text-align: right;
        }}
        .footer-name {{
            font-size: 14px;
            color: #1e3a5f;
            font-weight: 600;
        }}
        .footer-phone {{
            font-size: 18px;
            color: #c9a962;
            font-weight: 700;
            margin-top: 4px;
        }}
    </style>
</head>
<body>
    <div id="container">
        <div class="header">
            <div class="header-left">
                <div class="header-brand">{brand}</div>
                <div class="header-title">양도물건 소개</div>
            </div>
            <div class="header-right">
                <div class="header-date">{today}</div>
            </div>
        </div>
        
        <div class="license-section">{업종_html}</div>
        
        {"<div class='analysis-section'>" + pros_html + cons_html + "</div>" if pros_html or cons_html else ""}
        
        <div class="info-grid">
            <div class="info-item"><div class="info-label">회사형태</div><div class="info-value">{clean(data.get("회사형태"))}</div></div>
            <div class="info-item"><div class="info-label">자본금</div><div class="info-value">{clean(data.get("자본금"))}</div></div>
            <div class="info-item"><div class="info-label">소재지</div><div class="info-value">{clean(data.get("소재지"))}</div></div>
            <div class="info-item"><div class="info-label">법인설립</div><div class="info-value">{clean(data.get("법인설립일"))}</div></div>
            <div class="info-item"><div class="info-label">협회</div><div class="info-value">{clean(data.get("협회가입"))}</div></div>
            <div class="info-item"><div class="info-label">출자좌수</div><div class="info-value">{출자좌수}</div></div>
            <div class="info-item"><div class="info-label">공제조합잔액</div><div class="info-value highlight">{잔액}</div></div>
            <div class="info-item"></div>
        </div>

        
        <div class="table-section">
            <div class="table-title">
                최근년도 매출실적
                <span class="table-unit">(단위: 억)</span>
            </div>
            <table>
                <thead>
                    <tr>
                        <th style="width:60px">업종</th>
                        <th style="width:50px">면허</th>
                        <th style="width:50px">시평</th>
                        <th>2020</th><th>2021</th><th>2022</th><th>2023</th><th>2024</th><th>2025</th>
                        <th style="width:60px">3년</th>
                        <th style="width:60px">5년</th>
                    </tr>
                </thead>
                <tbody>{table_rows}</tbody>
            </table>
        </div>
        
        {"<div class='admin-section'><div class='section-header'>⚠ 행정사항</div><ul>" + 행정_html + "</ul></div>" if 행정_html else ""}
        
        {"<div class='check-section'><div class='check-header'>📋 주요 체크사항</div><div class='check-items'>" + 비고_html + "</div></div>" if 비고_html else ""}
        
        <div class="price-section">
            <div class="price-label">양 도 가</div>
            <div class="price-value"><span class="price-won">₩</span>{price}</div>
        </div>
        
        <div class="footer">
            <div class="footer-left">
                <div class="footer-brand">{brand}</div>
                <div class="footer-site">{site}</div>
            </div>
            <div class="footer-contact">
                <div class="footer-name">{consultant}</div>
                <div class="footer-phone">{phone}</div>
            </div>
        </div>
    </div>
</body>
</html>'''
        
        temp_html = "temp_gabji.html"
        with open(temp_html, "w", encoding="utf-8") as f:
            f.write(html)
        
        if output_path is None:
            output_path = _default_gabji_output_path("png")
        
        driver = None
        try:
            opts = Options()
            opts.add_argument("--headless")
            opts.add_argument("--window-size=900,1400")
            opts.add_argument("--hide-scrollbars")
            opts.add_argument("--force-device-scale-factor=2")
            
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
            driver.get(f"file://{os.path.abspath(temp_html)}")
            time.sleep(1.5)
            driver.find_element(By.ID, "container").screenshot(output_path)
            
            print(f"? PNG ??: {output_path}")
            return output_path
        except Exception as e:
            print(f"? PNG ?? ??: {e}")
            return None
        finally:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    pass
            if os.path.exists(temp_html):
                try:
                    os.remove(temp_html)
                except OSError:
                    pass


# =================================================================
# [Output] HTML
# =================================================================
class HTMLGenerator:
    def generate(self, data: dict, yangdo_price: str = None, output_path: str = None) -> str:
        if output_path is None:
            output_path = _default_gabji_output_path("html")
        
        png_gen = PNGGenerator()
        temp_html = "temp_view.html"
        
        # PNG 생성기의 HTML 구조를 재사용하되 파일로 저장
        today = datetime.now().strftime("%Y. %m. %d.")
        price = yangdo_price if yangdo_price else data.get("양도가", "협의")
        if not price or not str(price).strip(): price = "협의"
        
        업종 = " ".join([f'<span style="background:#1e3a5f;color:#fff;padding:8px 16px;margin:4px;border-radius:4px;display:inline-block;font-weight:600;">{_format_license_name(i.get("업종",""))}</span>' for i in data.get("업종정보", [])])
        
        비고 = data.get("비고", [])
        if isinstance(비고, str): 비고 = [비고]
        비고_html = " | ".join(비고) if 비고 else "-"
        
        html = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>갑지 - {CONFIG["BRAND_NAME"]}</title>
<style>body{{font-family:'Pretendard',sans-serif;background:#f5f5f5;padding:40px;}}
.c{{max-width:800px;margin:0 auto;background:#fff;box-shadow:0 4px 20px rgba(0,0,0,0.1);}}</style>
</head><body><div class="c">
<div style="background:#1e3a5f;color:#fff;padding:28px 40px;display:flex;justify-content:space-between;align-items:center;">
<div><div style="font-size:11px;color:#c9a962;letter-spacing:2px;margin-bottom:6px;">{CONFIG["BRAND_NAME"]}</div>
<div style="font-size:20px;font-weight:600;">양도물건 소개</div></div>
<div style="font-size:13px;color:rgba(255,255,255,0.6);">{today}</div></div>
<div style="padding:20px 40px;background:#f8f9fb;">{업종}</div>
<div style="padding:20px 40px;">
<p><b>회사형태:</b> {data.get("회사형태","-")} | <b>자본금:</b> {data.get("자본금","-")} | <b>소재지:</b> {data.get("소재지","-")}</p>
<p><b>법인설립:</b> {data.get("법인설립일","-")} | <b>협회:</b> {data.get("협회가입","-")} | <b>공제조합잔액:</b> <span style="color:#d32f2f;font-weight:700;">{data.get("공제조합잔액","-")}</span></p>
</div>
<div style="padding:16px 40px;background:#e3f2fd;"><b style="color:#1565c0;">📋 체크사항:</b> {비고_html}</div>
<div style="background:#1e3a5f;padding:32px;text-align:center;">
<div style="color:#c9a962;font-size:12px;letter-spacing:3px;margin-bottom:8px;">양도가</div>
<div style="color:#fff;font-size:40px;font-weight:300;"><span style="color:#c9a962;font-size:20px;margin-right:6px;">₩</span>{price}</div>
</div>
<div style="padding:20px 40px;display:flex;justify-content:space-between;background:#fafafa;border-top:1px solid #eee;">
<div><b style="color:#1e3a5f;">{CONFIG["BRAND_NAME"]}</b><br><span style="color:#888;font-size:12px;">{CONFIG["MAIN_SITE"]}</span></div>
<div style="text-align:right;"><span style="color:#1e3a5f;">{CONFIG["CONSULTANT_NAME"]}</span><br><span style="color:#c9a962;font-weight:700;font-size:16px;">{CONFIG["PHONE"]}</span></div>
</div></div></body></html>'''
        
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        
        print(f"✅ HTML 저장: {output_path}")
        return output_path


# =================================================================
# [Output] Excel
# =================================================================
class ExcelGenerator:
    def generate(self, data: dict, yangdo_price: str = None, output_path: str = None) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("❌ pip install openpyxl")
            return None
        
        if output_path is None:
            output_path = _default_gabji_output_path("xlsx")
        
        price = yangdo_price if yangdo_price else data.get("양도가", "협의")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "양도물건"
        
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws.merge_cells('A1:K1')
        ws['A1'] = f"{CONFIG['BRAND_NAME']} 양도물건 소개"
        ws['A1'].font = Font(bold=True, size=14, color="1E3A5F")
        
        row = 3
        for k, v in [("회사형태", data.get("회사형태")), ("자본금", data.get("자본금")), 
                     ("소재지", data.get("소재지")), ("공제조합잔액", data.get("공제조합잔액"))]:
            ws[f'A{row}'], ws[f'B{row}'] = k, v
            row += 1
        
        row += 1
        for col, h in enumerate(["업종","면허","시평","2020","2021","2022","2023","2024","2025","3년","5년"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font, cell.fill = header_font, header_fill
        
        for info in data.get("업종정보", []):
            row += 1
            매출 = info.get("매출", {})
            for col, v in enumerate([info.get("업종"), info.get("면허년도"), 
                info.get("시공능력평가액") or info.get("시평"),
                매출.get("2020년"), 매출.get("2021년"), 매출.get("2022년"),
                매출.get("2023년"), 매출.get("2024년"), 매출.get("2025년"),
                info.get("3년합계"), info.get("5년합계")], 1):
                ws.cell(row=row, column=col, value=v)
        
        row += 2
        ws[f'A{row}'] = f"양도가: {price}"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1E3A5F")
        
        wb.save(output_path)
        print(f"✅ Excel 저장: {output_path}")
        return output_path


# =================================================================
# [Output] 텍스트
# =================================================================
class TextGenerator:
    def generate(self, data: dict, yangdo_price: str = None) -> str:
        price = yangdo_price if yangdo_price else data.get("양도가", "협의")
        if not price: price = "협의"
        
        업종목록 = []
        seen_license = set()
        for info in data.get("업종정보", []):
            name = str(info.get("업종", "")).strip()
            if not name:
                continue
            key = re.sub(r"\s+", "", name.lower())
            if key in seen_license:
                continue
            seen_license.add(key)
            업종목록.append(_format_license_name(name))
        if not 업종목록:
            업종목록 = ["면허정보 확인중"]
        업종_text = "\n".join([f"  ◆ {idx}. {name}" for idx, name in enumerate(업종목록, start=1)])
        
        비고 = data.get("비고", [])
        if isinstance(비고, str): 비고 = [비고]
        비고_text = "\n  ".join([f"✓ {i}" for i in 비고 if i])
        
        return f"""━━━━━━━━━━━━━━━━━━━━━━━━━━━
{CONFIG["BRAND_NAME"]} 양도물건 소개
{datetime.now().strftime("%Y. %m. %d.")}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

[보유 면허]
{업종_text}

회사형태: {data.get('회사형태','-')}
자본금: {data.get('자본금','-')}
소재지: {data.get('소재지','-')}
법인설립: {data.get('법인설립일','-')}
협회: {data.get('협회가입','-')}
공제조합잔액: {data.get('공제조합잔액','-')}

[체크사항]
  {비고_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━
양도가: {price}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

{CONFIG["BRAND_NAME"]} ({CONFIG["MAIN_SITE"]})
{CONFIG["CONSULTANT_NAME"]} {CONFIG["PHONE"]}"""


# =================================================================
# [GUI]
# =================================================================
class GabjiApp:
    def __init__(self):
        self.generator = None
        self.sheet_lookup = None
        self.extracted_data = None
        
        self.root = tk.Tk()
        self.root.title("갑지 생성기 v3.0")
        self.root.geometry("900x750")
        self.root.configure(bg="#1e3a5f")
        
        self._create_widgets()
    
    def _create_widgets(self):
        # 헤더
        header = tk.Frame(self.root, bg="#1e3a5f", pady=12)
        header.pack(fill="x")
        tk.Label(header, text="📋 갑지 자동 생성기", font=("맑은 고딕", 18, "bold"),
                 fg="white", bg="#1e3a5f").pack()
        tk.Label(header, text="건설업 양도매물 정보 자동 변환", font=("맑은 고딕", 10),
                 fg="#c9a962", bg="#1e3a5f").pack()
        
        # 입력 영역
        input_frame = tk.Frame(self.root, bg="#0d2137", pady=12)
        input_frame.pack(fill="x")

        tk.Label(input_frame, text="등록번호:", font=("맑은 고딕", 10), fg="white", bg="#0d2137").pack(side="left", padx=(20, 5))
        self.reg_entry = tk.Entry(input_frame, font=("맑은 고딕", 11), width=12)
        self.reg_entry.pack(side="left")
        tk.Button(
            input_frame,
            text="🔎 번호 조회",
            font=("맑은 고딕", 10, "bold"),
            bg="#ffffff",
            fg="#1e3a5f",
            padx=12,
            pady=6,
            command=self._load_by_registration,
        ).pack(side="left", padx=(8, 14))

        tk.Button(input_frame, text="📁 이미지 선택", font=("맑은 고딕", 11, "bold"),
                  bg="#c9a962", fg="#1e3a5f", padx=20, pady=6,
                  command=self._select_image).pack(side="left", padx=(20, 15))
        
        tk.Label(input_frame, text="양도가:", font=("맑은 고딕", 10),
                 fg="white", bg="#0d2137").pack(side="left")
        
        self.price_entry = tk.Entry(input_frame, font=("맑은 고딕", 11), width=12)
        self.price_entry.pack(side="left", padx=(5, 15))
        self.price_entry.insert(0, "협의")
        
        self.status_label = tk.Label(input_frame, text="", font=("맑은 고딕", 9),
                                      fg="#c9a962", bg="#0d2137")
        self.status_label.pack(side="right", padx=20)
        
        # 출력 버튼
        btn_frame = tk.Frame(self.root, bg="#152d45", pady=10)
        btn_frame.pack(fill="x")
        
        tk.Label(btn_frame, text="출력:", font=("맑은 고딕", 9),
                 fg="#c9a962", bg="#152d45").pack(side="left", padx=(20, 8))
        
        self.png_btn = tk.Button(btn_frame, text="🖼 PNG", font=("맑은 고딕", 9),
                                  bg="white", fg="#1e3a5f", padx=10, pady=4,
                                  command=self._save_png, state="disabled")
        self.png_btn.pack(side="left", padx=4)
        
        self.html_btn = tk.Button(btn_frame, text="🌐 HTML", font=("맑은 고딕", 9),
                                   bg="white", fg="#1e3a5f", padx=10, pady=4,
                                   command=self._save_html, state="disabled")
        self.html_btn.pack(side="left", padx=4)
        
        self.excel_btn = tk.Button(btn_frame, text="📊 Excel", font=("맑은 고딕", 9),
                                    bg="white", fg="#1e3a5f", padx=10, pady=4,
                                    command=self._save_excel, state="disabled")
        self.excel_btn.pack(side="left", padx=4)
        
        self.copy_btn = tk.Button(btn_frame, text="📋 복사", font=("맑은 고딕", 9),
                                   bg="white", fg="#1e3a5f", padx=10, pady=4,
                                   command=self._copy_text, state="disabled")
        self.copy_btn.pack(side="left", padx=4)
        
        # 결과 영역
        result_frame = tk.Frame(self.root, bg="#1e3a5f", padx=16, pady=8)
        result_frame.pack(fill="both", expand=True)
        
        self.result_text = scrolledtext.ScrolledText(result_frame, font=("Consolas", 10),
                                                      bg="#0d1b2a", fg="#e0e0e0", wrap="word")
        self.result_text.pack(fill="both", expand=True)
        self.result_text.insert("1.0", "\n  이미지를 선택하거나 등록번호를 입력해 시트에서 불러오세요.\n\n  [새로운 기능]\n  • 등록번호 기반 시트 조회 생성\n  • 청구 양도가 범위의 마지막 값 자동 반영\n  • 매물 장단점 자동 분석\n  • 프리미엄 디자인 출력")
        
        # 푸터
        footer = tk.Frame(self.root, bg="#0d2137", pady=6)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text=f"{CONFIG['BRAND_NAME']} | {CONFIG['PHONE']}",
                 font=("맑은 고딕", 9), fg="#c9a962", bg="#0d2137").pack()
    
    def _get_price(self):
        raw = self.price_entry.get().strip()
        normalized = extract_final_yangdo_price(raw)
        if normalized != raw:
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, normalized)
        return normalized or "협의"

    def _load_by_registration(self):
        registration_no = self.reg_entry.get().strip()
        if not registration_no:
            messagebox.showwarning("입력 필요", "등록번호를 입력해주세요.")
            return

        self.status_label.config(text="시트 조회 중...")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", f"🔎 등록번호 '{registration_no}' 조회 중...\n")
        self.root.update()

        try:
            if self.sheet_lookup is None:
                self.sheet_lookup = ListingSheetLookup()
            data = self.sheet_lookup.load_listing(registration_no)
            source_label = f"시트 조회 / 등록번호 {registration_no}"
            source_uid = str(data.get("원본UID", "")).strip()
            if source_uid:
                source_label = f"{source_label} / 원본동기화 UID {source_uid}"
            self._apply_extracted_data(data, source_label=source_label)
        except Exception as e:
            self.status_label.config(text="조회 실패")
            self.result_text.delete("1.0", "end")
            self.result_text.insert("1.0", f"❌ 시트 조회 오류: {e}")
    
    def _select_image(self):
        filepath = filedialog.askopenfilename(
            title="이미지 선택",
            filetypes=[("이미지", "*.png *.jpg *.jpeg *.gif *.webp"), ("모든 파일", "*.*")]
        )
        if filepath: self._process_image(filepath)

    def _apply_extracted_data(self, data, source_label=""):
        self.extracted_data = data or {}

        reg = str(self.extracted_data.get("등록번호", "")).strip()
        if reg:
            self.reg_entry.delete(0, tk.END)
            self.reg_entry.insert(0, reg)

        ep = extract_final_yangdo_price(self.extracted_data.get("양도가", ""))
        if ep:
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, ep)

        analysis = analyze_property(self.extracted_data)
        text = TextGenerator().generate(self.extracted_data, self._get_price())
        if source_label:
            text = f"[데이터 소스] {source_label}\n\n{text}"

        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", text)

        if analysis["장점"] or analysis["단점"]:
            self.result_text.insert("end", "\n\n[자동 분석 결과]")
            if analysis["장점"]:
                self.result_text.insert("end", f"\n  👍 장점: {', '.join(analysis['장점'])}")
            if analysis["단점"]:
                self.result_text.insert("end", f"\n  ⚠️ 주의: {', '.join(analysis['단점'])}")

        for btn in [self.png_btn, self.html_btn, self.excel_btn, self.copy_btn]:
            btn.config(state="normal")
        self.status_label.config(text="✅ 완료")
    
    def _process_image(self, filepath):
        self.status_label.config(text="분석 중...")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", "🔍 이미지 분석 중...\n")
        self.root.update()
        
        try:
            if self.generator is None:
                self.generator = GabjiGenerator()
            self.extracted_data = self.generator.analyze_image(filepath)
            
            if self.extracted_data:
                self._apply_extracted_data(self.extracted_data, source_label="이미지 AI 분석")
            else:
                self.result_text.insert("1.0", "❌ 분석 실패")
                self.status_label.config(text="실패")
        except Exception as e:
            self.result_text.insert("1.0", f"❌ 오류: {e}")
    
    def _save_png(self):
        if self.extracted_data:
            self.status_label.config(text="PNG 생성 중...")
            self.root.update()
            path = PNGGenerator().generate(self.extracted_data, self._get_price())
            if path:
                self.status_label.config(text="✅ PNG 저장됨")
                messagebox.showinfo("완료", f"저장됨:\n{path}")
    
    def _save_html(self):
        if self.extracted_data:
            path = HTMLGenerator().generate(self.extracted_data, self._get_price())
            if path:
                self.status_label.config(text="✅ HTML 저장됨")
                os.startfile(path)
    
    def _save_excel(self):
        if self.extracted_data:
            path = ExcelGenerator().generate(self.extracted_data, self._get_price())
            if path:
                self.status_label.config(text="✅ Excel 저장됨")
                messagebox.showinfo("완료", f"저장됨:\n{path}")
    
    def _copy_text(self):
        if self.extracted_data:
            text = TextGenerator().generate(self.extracted_data, self._get_price())
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.status_label.config(text="📋 복사됨")
    
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    try:
        GabjiApp().run()
    except ValueError as e:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("환경 설정 오류", str(e))
        root.destroy()
        raise SystemExit(1)
