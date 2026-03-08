#!/usr/bin/env python3
"""
=============================================================
  건설업 수주역량 자동진단 시스템 v3.0
  행정사사무소 하랑 · 서울건설정보
=============================================================

  [1] 나라장터 발주계획 자동수집 + 업종분류 + 재량계약 필터
  [2] 금융위 기업재무정보 → 고객사 재무 자동조회
  [3] 고객사 × 발주계획 매칭
  [4] 수주역량 진단 리포트 (엑셀)

  ★ config.txt에 키만 넣으면 나머지는 전자동
  ★ run.bat 더블클릭으로 실행
=============================================================
"""

import os, sys, json, time, logging, argparse, csv, sqlite3, shutil
from datetime import datetime, timedelta
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.parse import urlencode
from urllib.error import URLError, HTTPError

# ── 경로 ──
BASE_DIR = Path(__file__).parent.resolve()
CONFIG_FILE = BASE_DIR / "config.txt"
OUTPUT_DIR = BASE_DIR / "result"
LOG_DIR = BASE_DIR / "logs"
G2B_PROFILE_CACHE_FILE = LOG_DIR / "g2b_api_profile_cache.json"
RAW_LOG_DIR = LOG_DIR / "raw"
CLIENTS_CSV_FILE = BASE_DIR / "clients.csv"
SQLITE_DB_FILE = OUTPUT_DIR / "g2b_history.db"
LATEST_XLSX_FILE = OUTPUT_DIR / "latest.xlsx"

LOG_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
RAW_LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(LOG_DIR / f"log_{datetime.now():%Y%m%d}.txt"), encoding="utf-8"),
    ]
)
log = logging.getLogger()


# =============================================================
#  설정
# =============================================================

DEFAULT_CONFIG = """# ============================================
#  건설업 수주역량 자동진단 설정파일
#  행정사사무소 하랑 · 서울건설정보
# ============================================
#
# ★ 공공데이터포털 마이페이지 > 인증키에서
#   Encoding 키와 Decoding 키 둘 다 넣으세요
#   (프로그램이 자동으로 맞는 키를 찾습니다)
# ★ 환경변수 ENCODING_KEY / DECODING_KEY를 설정하면
#   config.txt보다 환경변수를 우선 사용합니다.
#

ENCODING_KEY=여기에_인코딩키_붙여넣기
DECODING_KEY=여기에_디코딩키_붙여넣기

# ── 아래는 수정 안해도 됩니다 ──
YEAR={year}
DISC_LIMIT=20000
SMALL_LIMIT=5000
LOG_RETENTION_DAYS=30
WEBHOOK_URL=

# ── 인큐베이팅 고객사 ──
# 형식: 회사명:업종:법인등록번호
# 법인등록번호가 있으면 재무정보를 자동 조회합니다
# 법인등록번호가 없으면 :만 쓰거나 비워두세요
#
# 예시:
# CLIENTS=하늘건설:일반건설:1101111234567,동방전기:전기공사:1301110001234,서울ICT:정보통신:
CLIENTS=
"""


def load_config():
    if not CONFIG_FILE.exists():
        txt = DEFAULT_CONFIG.format(year=datetime.now().year)
        CONFIG_FILE.write_text(txt, encoding="utf-8-sig")
        return None

    config = {}
    for line in CONFIG_FILE.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, v = line.split("=", 1)
            config[k.strip()] = v.strip()
    return config


def parse_config_clients(raw):
    clients = []
    if not raw:
        return clients
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        segs = part.split(":")
        name = segs[0].strip() if len(segs) > 0 else ""
        cat = segs[1].strip() if len(segs) > 1 else ""
        crno = segs[2].strip() if len(segs) > 2 else ""
        if name:
            clients.append({"name": name, "category": cat, "crno": crno})
    return clients


def load_clients_csv():
    """clients.csv 로드 (name, category, crno)."""
    if not CLIENTS_CSV_FILE.exists():
        return []
    clients = []
    try:
        with CLIENTS_CSV_FILE.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("name") or row.get("회사명") or "").strip()
                category = (row.get("category") or row.get("업종") or "").strip()
                crno = (row.get("crno") or row.get("법인등록번호") or "").strip()
                if name:
                    clients.append({"name": name, "category": category, "crno": crno})
    except Exception as e:
        log.warning(f"  clients.csv 로드 실패: {e}")
    return clients


def ensure_clients_csv_template():
    if CLIENTS_CSV_FILE.exists():
        return
    try:
        CLIENTS_CSV_FILE.write_text("name,category,crno\n", encoding="utf-8")
    except Exception:
        pass


def parse_clients(config):
    """고객사 파싱(config + clients.csv 병합) → [{name, category, crno}, ...]"""
    merged = {}
    for cl in parse_config_clients(config.get("CLIENTS", "")):
        merged[cl["name"]] = cl
    for cl in load_clients_csv():
        prev = merged.get(cl["name"], {})
        merged[cl["name"]] = {
            "name": cl["name"],
            "category": cl.get("category") or prev.get("category", ""),
            "crno": cl.get("crno") or prev.get("crno", ""),
        }
    return list(merged.values())


def resolve_target_year(config):
    """YEAR 설정값을 안전하게 정규화."""
    now_year = datetime.now().year
    raw_year = (config.get("YEAR", "") or "").strip()
    if not raw_year:
        return now_year
    try:
        year = int(raw_year)
    except ValueError:
        log.warning(f"  YEAR 설정값이 숫자가 아닙니다: {raw_year} (현재연도 {now_year}로 대체)")
        return now_year
    if year < 2000 or year > now_year + 2:
        log.warning(f"  YEAR 설정값 범위가 비정상입니다: {year} (현재연도 {now_year}로 대체)")
        return now_year
    return year


def resolve_api_keys(config):
    """환경변수 우선, 없으면 config.txt 키 사용."""
    env_enc = (os.environ.get("ENCODING_KEY", "") or "").strip()
    env_dec = (os.environ.get("DECODING_KEY", "") or "").strip()
    cfg_enc = (config.get("ENCODING_KEY", "") or "").strip()
    cfg_dec = (config.get("DECODING_KEY", "") or "").strip()

    enc = env_enc or cfg_enc
    dec = env_dec or cfg_dec
    sources = {
        "encoding": "env" if env_enc else "config",
        "decoding": "env" if env_dec else "config",
    }
    return enc, dec, sources


def load_g2b_profile_cache():
    if not G2B_PROFILE_CACHE_FILE.exists():
        return {}
    try:
        raw = json.loads(G2B_PROFILE_CACHE_FILE.read_text(encoding="utf-8"))
        return raw if isinstance(raw, dict) else {}
    except Exception as e:
        log.warning(f"  프로파일 캐시 로드 실패: {e}")
        return {}


def save_g2b_profile_cache(cache):
    try:
        G2B_PROFILE_CACHE_FILE.write_text(
            json.dumps(cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        log.warning(f"  프로파일 캐시 저장 실패: {e}")


def cleanup_old_logs(retention_days=30):
    """지정 일수보다 오래된 로그 파일 정리."""
    if retention_days <= 0:
        return
    cutoff = datetime.now() - timedelta(days=retention_days)
    removed = 0
    for root in [LOG_DIR, RAW_LOG_DIR]:
        for path in root.glob("*"):
            try:
                if path.is_file() and datetime.fromtimestamp(path.stat().st_mtime) < cutoff:
                    path.unlink()
                    removed += 1
            except Exception:
                continue
    if removed:
        log.info(f"  로그 정리 완료: {removed}개 파일 삭제 (보관 {retention_days}일)")


def mask_service_key(url):
    if "serviceKey=" not in url:
        return url
    head, tail = url.split("serviceKey=", 1)
    if "&" in tail:
        key, rest = tail.split("&", 1)
        return f"{head}serviceKey=***&{rest}"
    return f"{head}serviceKey=***"


def write_raw_sample(call_label, key_label, url, status, payload):
    """API 원시 응답 샘플 저장 (운영 장애 재현용)."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    safe_label = call_label.replace(":", "_").replace("/", "_")
    fpath = RAW_LOG_DIR / f"{ts}_{safe_label}_{key_label}_{status}.jsonl"
    body = payload if isinstance(payload, str) else str(payload)
    record = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "call_label": call_label,
        "key_label": key_label,
        "status": status,
        "url": mask_service_key(url),
        "payload_head": body[:20000],
    }
    try:
        fpath.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def wait_for_enter(message, interactive):
    """Interactive 모드에서만 입력 대기."""
    if not interactive:
        return
    try:
        input(message)
    except EOFError:
        log.warning("  비대화 환경으로 감지되어 입력 대기를 건너뜁니다.")


# =============================================================
#  공통 API 호출
# =============================================================

API_FAILURE_HINTS = {
    "missing-required-params-or-invalid-endpoint": "404 반복: 필수 조회 파라미터 누락 또는 엔드포인트 변경 가능성이 큽니다.",
    "auth-or-permission": "401/403 발생: 활용신청 승인 상태, 키 권한, 서비스 연결 상태를 점검하세요.",
    "upstream-server-error": "5xx 발생: 일시 장애일 수 있습니다. 잠시 후 재시도하세요.",
    "invalid-response-format": "응답 파싱 실패: XML/HTML 오류 응답 또는 응답 포맷 변경 가능성이 있습니다.",
    "network-error": "네트워크 오류: DNS/프록시/방화벽 설정을 확인하세요.",
}


def classify_api_failures(failures):
    if not failures:
        return "unknown"
    if all("HTTP404" in f for f in failures):
        return "missing-required-params-or-invalid-endpoint"
    if any("HTTP401" in f or "HTTP403" in f for f in failures):
        return "auth-or-permission"
    if any("HTTP500" in f or "HTTP502" in f or "HTTP503" in f for f in failures):
        return "upstream-server-error"
    if any("invalid-response" in f for f in failures):
        return "invalid-response-format"
    if any("URLError" in f or "TimeoutError" in f for f in failures):
        return "network-error"
    return "unknown"


def try_api_call(
    url_template,
    encoding_key,
    decoding_key,
    extra_params=None,
    call_label="API",
    raw_sample_mode="error",
):
    """
    인코딩키 → 디코딩키 순서로 시도.
    조달청은 인코딩키, 금융위는 디코딩키를 주로 씀.
    """
    failures = []
    for key_label, key_val in [("encoding", encoding_key), ("decoding", decoding_key)]:
        if not key_val or "여기에" in key_val:
            continue
        full_url = url_template
        raw = ""
        try:
            params = {"serviceKey": key_val}
            if extra_params:
                params.update(extra_params)

            # 조달청 API는 serviceKey를 직접 URL에 넣어야 함 (이미 인코딩된 키)
            if key_label == "encoding":
                # 인코딩키는 이미 URL인코딩 되어있으므로 직접 붙임
                param_str = f"serviceKey={key_val}"
                other_params = {k: v for k, v in params.items() if k != "serviceKey"}
                if other_params:
                    param_str += "&" + urlencode(other_params)
                full_url = f"{url_template}?{param_str}"
            else:
                # 디코딩키는 urlencode가 자동 인코딩
                full_url = f"{url_template}?{urlencode(params)}"

            req = Request(full_url, headers={"Accept": "application/json"})
            with urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8")
                status_code = getattr(resp, "status", 200)

            # XML 에러 응답 체크
            if raw.strip().startswith("<"):
                if "SERVICE_KEY_IS_NOT_REGISTERED_ERROR" in raw:
                    if raw_sample_mode in ("all", "error"):
                        write_raw_sample(call_label, key_label, full_url, "xml-error", raw)
                    continue
                if "UNREGISTERED_SERVICE_ERROR" in raw:
                    if raw_sample_mode in ("all", "error"):
                        write_raw_sample(call_label, key_label, full_url, "xml-error", raw)
                    continue
                if "<resultCode>00</resultCode>" not in raw and "<resultCode>0</resultCode>" not in raw:
                    if raw_sample_mode in ("all", "error"):
                        write_raw_sample(call_label, key_label, full_url, "xml-error", raw)
                    continue

            data = json.loads(raw)
            if raw_sample_mode in ("all", "success"):
                write_raw_sample(call_label, key_label, full_url, status_code, raw)
            return data, key_label, None

        except (json.JSONDecodeError, UnicodeDecodeError):
            if raw_sample_mode in ("all", "error"):
                write_raw_sample(call_label, key_label, full_url, "json-decode-error", raw if "raw" in locals() else "")
            failures.append(f"{key_label}:invalid-response")
            continue
        except HTTPError as e:
            body = ""
            try:
                body = e.read().decode("utf-8", errors="replace")
            except Exception:
                body = ""
            if raw_sample_mode in ("all", "error"):
                write_raw_sample(call_label, key_label, full_url, f"HTTP{e.code}", body)
            if e.code in (401, 403, 404, 500, 502, 503):
                log.warning(f"  HTTP {e.code} ({key_label}키) - 다음 키 시도")
                failures.append(f"{key_label}:HTTP{e.code}")
                continue
            failures.append(f"{key_label}:HTTP{e.code}")
            raise
        except Exception as e:
            if raw_sample_mode in ("all", "error"):
                write_raw_sample(call_label, key_label, full_url, type(e).__name__, str(e))
            failures.append(f"{key_label}:{type(e).__name__}")
            continue

    if not failures:
        return None, None, {"classification": "missing-key", "failures": []}

    uniq = []
    for f in failures:
        if f not in uniq:
            uniq.append(f)
    classification = classify_api_failures(uniq)
    log.error(f"  {call_label} 호출 실패: {', '.join(uniq)}")
    hint = API_FAILURE_HINTS.get(classification)
    if hint:
        log.error(f"  {hint}")
    return None, None, {"classification": classification, "failures": uniq}


# =============================================================
#  나라장터 발주계획 API
# =============================================================

G2B_BASE = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService"
G2B_EP_CANDIDATES = {
    "const": ["getOrderPlanSttusListCnstwk"],
    # 운영 중 표기 흔들림(servc/servce)이 있어 후보를 순차 시도
    "serv": ["getOrderPlanSttusListServc", "getOrderPlanSttusListServce"],
}
G2B_LABEL = {"const": "공사", "serv": "용역"}


def build_g2b_param_profiles(year):
    """조회 파라미터 프로파일 후보."""
    y = int(year)
    return [
        {
            "name": "order_ym",
            "static_params": {
                "inqryDiv": "1",
                "orderBgnYm": f"{y}01",
                "orderEndYm": f"{y}12",
            },
        },
        {
            "name": "date_range",
            "static_params": {
                "inqryDiv": "1",
                "inqryBgnDt": f"{y}0101",
                "inqryEndDt": f"{y}1231",
            },
        },
        {
            "name": "minimal",
            "static_params": {},
        },
    ]


def make_g2b_params(profile, page, rows):
    params = {"pageNo": str(page), "numOfRows": str(rows), "type": "json"}
    params.update(profile["static_params"])
    return params


def g2b_fetch_page(enc_key, dec_key, endpoint_name, params):
    url = f"{G2B_BASE}/{endpoint_name}"

    data, used_key, fail_info = try_api_call(
        url, enc_key, dec_key, params, call_label=f"G2B:{endpoint_name}", raw_sample_mode="all"
    )
    if not data:
        return [], 0, None, False, fail_info

    if "response" in data:
        header = data["response"].get("header", {})
        if header.get("resultCode") != "00":
            log.error(f"  G2B 오류: {header.get('resultMsg', '?')}")
            return [], 0, used_key, False, {
                "classification": "api-result-error",
                "failures": [header.get("resultMsg", "unknown-result-msg")],
            }
        body = data["response"].get("body", {})
        items = body.get("items", [])
        if isinstance(items, dict):
            items = items.get("item", [])
        if isinstance(items, dict):
            items = [items]
        return items, int(body.get("totalCount", 0)), used_key, True, None
    return [], 0, used_key, False, {
        "classification": "unexpected-payload",
        "failures": ["response-missing"],
    }


def summarize_failures(fail_infos):
    counter = {}
    for info in fail_infos:
        if not info:
            continue
        key = info.get("classification", "unknown")
        counter[key] = counter.get(key, 0) + 1
    return ", ".join(f"{k}:{v}" for k, v in sorted(counter.items())) if counter else "unknown"


def build_g2b_trials(endpoint_candidates, profiles, preferred):
    profile_by_name = {p["name"]: p for p in profiles}
    trials = []
    seen = set()

    if preferred and isinstance(preferred, dict):
        pref_endpoint = preferred.get("endpoint", "")
        pref_profile = preferred.get("profile", "")
        if pref_endpoint in endpoint_candidates and pref_profile in profile_by_name:
            trials.append((pref_endpoint, profile_by_name[pref_profile], True))
            seen.add((pref_endpoint, pref_profile))

    for endpoint_name in endpoint_candidates:
        for profile in profiles:
            key = (endpoint_name, profile["name"])
            if key in seen:
                continue
            trials.append((endpoint_name, profile, False))
    return trials


def g2b_fetch_all(enc_key, dec_key, ep_key, year, max_pages=30, preferred=None, rows=999):
    label = G2B_LABEL[ep_key]
    endpoint_candidates = G2B_EP_CANDIDATES[ep_key]
    profiles = build_g2b_param_profiles(year)
    trials = build_g2b_trials(endpoint_candidates, profiles, preferred)

    all_items = []
    total = 0
    working_key_type = None
    selected_endpoint = None
    selected_profile = None
    used_cached_trial = False
    page_error = None
    probe_failures = []

    # 1페이지를 기준으로 유효 endpoint/profile 조합 탐색
    for endpoint_name, profile, is_cached_trial in trials:
        params = make_g2b_params(profile, page=1, rows=rows)
        if is_cached_trial:
            log.info(f"  [{label}] preflight(cache) endpoint={endpoint_name}, profile={profile['name']}...")
        else:
            log.info(f"  [{label}] preflight endpoint={endpoint_name}, profile={profile['name']}...")
        items, total, kt, ok, fail_info = g2b_fetch_page(enc_key, dec_key, endpoint_name, params)
        if ok:
            selected_endpoint = endpoint_name
            selected_profile = profile
            used_cached_trial = is_cached_trial
            all_items.extend(items)
            if kt:
                working_key_type = kt
            log.info(
                f"  [{label}] 연결 성공 endpoint={selected_endpoint}, profile={selected_profile['name']}, key={working_key_type or '-'}"
            )
            break
        probe_failures.append(fail_info)

    if not selected_endpoint:
        log.error(
            f"  [{label}] 유효 호출 조합 탐색 실패 (endpoint/profile 전부 실패: {summarize_failures(probe_failures)})"
        )
        return [], {
            "selected_endpoint": None,
            "selected_profile": None,
            "used_cache": False,
            "fail_summary": summarize_failures(probe_failures),
            "page_error": None,
        }

    page = 2
    while page <= max_pages and len(all_items) < total:
        log.info(f"  [{label}] {page}페이지...")
        params = make_g2b_params(selected_profile, page=page, rows=rows)
        items, total_page, kt, ok, fail_info = g2b_fetch_page(enc_key, dec_key, selected_endpoint, params)
        if not ok:
            if fail_info:
                log.warning(
                    f"  [{label}] 페이지 수집 중단 page={page} ({fail_info.get('classification', 'unknown')})"
                )
                page_error = fail_info.get("classification", "unknown")
            break
        if kt and not working_key_type:
            working_key_type = kt
        if total_page:
            total = total_page
        if not items:
            break
        all_items.extend(items)
        page += 1
        time.sleep(0.5)

    log.info(
        f"  [{label}] {len(all_items)}건 수집 (전체 {total}건, endpoint={selected_endpoint}, profile={selected_profile['name']})"
    )
    return all_items, {
        "selected_endpoint": selected_endpoint,
        "selected_profile": selected_profile["name"],
        "used_cache": used_cached_trial,
        "fail_summary": summarize_failures(probe_failures),
        "page_error": page_error,
    }


# =============================================================
#  금융위 기업 재무정보 API
# =============================================================

FSC_BASE = "https://apis.data.go.kr/1160100/service/GetFinaStatInfoService_V2"
FSC_EPS = {
    "summary": "getSummFinaStat_V2",      # 요약재무제표
    "bs": "getBs_V2",                      # 재무상태표
    "income": "getIncoStat_V2",            # 손익계산서
}


def fsc_fetch(enc_key, dec_key, endpoint, crno="", corp_name="", biz_year=""):
    """금융위 재무정보 조회"""
    url = f"{FSC_BASE}/{FSC_EPS[endpoint]}"
    params = {
        "pageNo": "1",
        "numOfRows": "20",
        "resultType": "json",
    }
    if crno:
        params["crno"] = crno
    if corp_name:
        params["fnccmpNm"] = corp_name
    if biz_year:
        params["bizYear"] = biz_year

    data, used_key, _ = try_api_call(
        url, enc_key, dec_key, params, call_label=f"FSC:{FSC_EPS[endpoint]}"
    )
    if not data:
        return []

    try:
        items = data.get("response", {}).get("body", {}).get("items", {}).get("item", [])
        if isinstance(items, dict):
            items = [items]
        return items
    except (AttributeError, TypeError):
        return []


def get_client_financials(enc_key, dec_key, client):
    """고객사 1곳의 재무 요약 조회"""
    crno = client.get("crno", "")
    name = client.get("name", "")

    if not crno and not name:
        return None

    # 최근 2개년 시도
    current_year = datetime.now().year
    results = []

    for year in [current_year - 1, current_year - 2]:
        items = fsc_fetch(enc_key, dec_key, "summary",
                         crno=crno, corp_name=name if not crno else "",
                         biz_year=str(year))
        if items:
            results.extend(items)

    if not results:
        # 연도 없이 재시도
        items = fsc_fetch(enc_key, dec_key, "summary",
                         crno=crno, corp_name=name if not crno else "")
        if items:
            results = items

    if not results:
        return None

    # 가장 최신 데이터 선택 (개별재무제표 우선)
    best = None
    for item in results:
        dcd = item.get("fnclDcdNm", "")
        if "개별" in dcd or "별도" in dcd:
            if not best or item.get("bizYear", "") > best.get("bizYear", ""):
                best = item
    if not best and results:
        best = max(results, key=lambda x: x.get("bizYear", ""))

    if not best:
        return None

    def safe_int(v):
        try:
            return int(v) if v else 0
        except (ValueError, TypeError):
            return 0

    return {
        "기준연도": best.get("bizYear", ""),
        "재무제표구분": best.get("fnclDcdNm", ""),
        "매출액": safe_int(best.get("enpSaleAmt", 0)),
        "영업이익": safe_int(best.get("enpBzopPft", 0)),
        "당기순이익": safe_int(best.get("iclsPalClcAmt", 0)),
        "자산총계": safe_int(best.get("enpTastAmt", 0)),
        "부채총계": safe_int(best.get("enpTdbtAmt", 0)),
        "자본총계": safe_int(best.get("enpTcptlAmt", 0)),
    }


# =============================================================
#  데이터 처리
# =============================================================

CAT_RULES = {
    "전기공사": ["전기", "전력", "배전", "수변전", "조명", "전기설비", "전기공사"],
    "정보통신공사": ["정보통신", "통신", "네트워크", "CCTV", "방송", "통합배선", "ICT", "cctv"],
    "소방공사": ["소방", "방재", "소화", "경보", "스프링클러", "소방설비"],
    "기계설비공사": ["기계설비", "냉난방", "공조", "배관", "급배수"],
    "조경공사": ["조경", "녹화", "수목", "식재"],
    "건축공사": ["건축", "신축", "증축", "리모델링", "개보수", "인테리어"],
    "토목공사": ["토목", "도로", "상하수도", "하천", "교량", "포장"],
}


def classify(name):
    if not name:
        return "기타"
    u = name.upper()
    for cat, kws in CAT_RULES.items():
        for kw in kws:
            if kw.upper() in u:
                return cat
    return "기타"


def contract_type(amt, disc_limit, small_limit):
    if amt <= 0:
        return "금액미정"
    if amt <= small_limit:
        return "소액수의"
    if amt <= disc_limit:
        return "재량계약"
    return "일반경쟁"


def process_g2b(raw_items, config):
    disc = int(config.get("DISC_LIMIT", "20000")) * 10000
    small = int(config.get("SMALL_LIMIT", "5000")) * 10000

    results = []
    for item in raw_items:
        # 차세대 API: sumOrderAmt, 구버전: asignBdgtAmt
        try:
            amt = int(item.get("sumOrderAmt", 0) or item.get("asignBdgtAmt", 0) or 0)
        except (ValueError, TypeError):
            amt = 0

        # 차세대 API: bizNm, 구버전: bidNtceNm
        name = item.get("bizNm", "") or item.get("bidNtceNm", "") or item.get("prdctClsfcNoNm", "") or ""
        results.append({
            "번호": item.get("orderPlanUntyNo", ""),
            "사업명": name,
            "발주기관": item.get("orderInsttNm", ""),
            "수요기관": item.get("totlmngInsttNm", "") or item.get("dminsttNm", ""),
            "예산액": amt,
            "예산(만원)": round(amt / 10000) if amt > 0 else 0,
            "예산(억)": round(amt / 1_0000_0000, 2) if amt > 0 else 0,
            "발주시기": item.get("nticeDt", "") or item.get("orderPlanRegDt", ""),
            "계약방법": item.get("cntrctMthdNm", ""),
            "업종": classify(name),
            "재량여부": contract_type(amt, disc, small),
            "지역": item.get("cnstwkRgnNm", "") or item.get("dminsttRgnNm", "") or item.get("orderInsttRgnNm", ""),
            "담당자": item.get("ofclNm", "") or item.get("orderInsttOfclNm", ""),
            "연락처": item.get("telNo", "") or item.get("orderInsttOfclTelNo", ""),
        })
    return results


def compute_match_score(item, client_category):
    """단순 휴리스틱 매칭 점수(0~100)."""
    score = 60
    reasons = []

    if item.get("재량여부") == "재량계약":
        score += 15
        reasons.append("재량계약")
    elif item.get("재량여부") == "소액수의":
        score += 10
        reasons.append("소액수의")

    if item.get("연락처"):
        score += 10
        reasons.append("연락처 확보")
    if item.get("담당자"):
        score += 5
        reasons.append("담당자 확보")
    if item.get("지역"):
        score += 5
        reasons.append("지역 정보")

    # 고객 업종 키워드와 사업 업종이 유사하면 가점
    cc = client_category or ""
    ic = item.get("업종", "")
    if any(k in cc and k in ic for k in ["전기", "통신", "소방", "기계", "조경", "건축", "토목"]):
        score += 5
        reasons.append("업종 직접일치")

    return min(score, 100), ", ".join(reasons)


def match_clients(disc_data, clients):
    """고객사 × 재량계약 발주 매칭."""
    matched = []
    for cl in clients:
        cc = cl["category"]
        for d in disc_data:
            ic = d["업종"]
            ok = False
            if ("일반" in cc or "건축" in cc) and ("건축" in ic or "토목" in ic):
                ok = True
            elif "전기" in cc and "전기" in ic:
                ok = True
            elif "통신" in cc and "통신" in ic:
                ok = True
            elif "소방" in cc and "소방" in ic:
                ok = True
            elif "기계" in cc and "기계" in ic:
                ok = True
            elif "조경" in cc and "조경" in ic:
                ok = True
            if ok:
                score, reason = compute_match_score(d, cc)
                matched.append(
                    {
                        **d,
                        "매칭고객사": cl["name"],
                        "고객업종": cl["category"],
                        "매칭점수": score,
                        "매칭근거": reason,
                    }
                )
    matched.sort(key=lambda x: x.get("매칭점수", 0), reverse=True)
    return matched


# =============================================================
#  엑셀 리포트 생성
# =============================================================

def ensure_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        log.info("  openpyxl 설치중...")
        os.system(f'"{sys.executable}" -m pip install openpyxl -q')
        try:
            import openpyxl
            return True
        except ImportError:
            return False


def create_report(
    all_data,
    disc_data,
    matched,
    clients,
    financials,
    config,
    data_quality=None,
    runtime_meta=None,
    delta_info=None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # 스타일
    HF = PatternFill("solid", fgColor="1F4E79")
    HN = Font(bold=True, color="FFFFFF", size=10, name="Malgun Gothic")
    DF = Font(size=10, name="Malgun Gothic")
    TF = Font(bold=True, size=14, name="Malgun Gothic", color="1F4E79")
    SF = Font(bold=True, size=12, name="Malgun Gothic", color="1F4E79")
    AF = PatternFill("solid", fgColor="F2F7FB")
    GF = PatternFill("solid", fgColor="E2EFDA")
    WF = PatternFill("solid", fgColor="FFF2CC")
    BD = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def hdr(ws, row, n):
        for c in range(1, n + 1):
            cl = ws.cell(row=row, column=c)
            cl.fill, cl.font, cl.border = HF, HN, BD
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def aw(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 3, 10), 42)

    def fmt_money(v):
        """원 → 만원 표시"""
        if v == 0:
            return "-"
        return f"{round(v / 10000):,}"

    def fmt_eok(v):
        """원 → 억원 표시"""
        if v == 0:
            return "-"
        return f"{v / 1_0000_0000:,.1f}"

    # ═══════════════════════════════════════════
    # 시트1: 수주역량 진단 (고객사별 요약)
    # ═══════════════════════════════════════════
    ws0 = wb.active
    ws0.title = "수주역량 진단"
    ws0.sheet_properties.tabColor = "0F4C75"

    ws0["A1"] = "건설업 수주역량 진단 리포트"
    ws0["A1"].font = TF
    ws0.merge_cells("A1:H1")
    report_year = (runtime_meta or {}).get("target_year", config.get("YEAR", ""))
    ws0["A2"] = f"행정사사무소 하랑 | {datetime.now():%Y-%m-%d %H:%M} | {report_year}년"
    ws0["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")
    ws0.merge_cells("A2:H2")

    quality = data_quality or {"level": "UNKNOWN", "detail": "-"}
    level = quality.get("level", "UNKNOWN")
    detail = quality.get("detail", "")
    dq_color = {"FULL": "2E7D32", "PARTIAL": "B8860B", "FAIL": "C00000"}.get(level, "666666")
    ws0["A3"] = f"DATA_QUALITY: {level}" + (f" | {detail}" if detail else "")
    ws0["A3"].font = Font(bold=True, color="FFFFFF", size=10, name="Malgun Gothic")
    ws0["A3"].fill = PatternFill("solid", fgColor=dq_color)
    ws0["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws0.merge_cells("A3:H3")

    meta = runtime_meta or {}
    ws0["A4"] = (
        f"META: RUN_MODE={meta.get('run_mode', '-')}"
        f" | KEY_SOURCE={meta.get('key_source', '-')}"
        f" | PROFILE={meta.get('profile_info', '-')}"
    )
    ws0["A4"].font = Font(size=9, color="666666", name="Malgun Gothic")
    ws0.merge_cells("A4:H4")

    start_row = 6

    if clients:
        row = start_row
        for cl in clients:
            cn = cl["name"]
            cc = cl["category"]
            fin = financials.get(cn)
            mc = [m for m in matched if m["매칭고객사"] == cn]

            # 고객사 헤더
            ws0.cell(row=row, column=1, value=f"■ {cn}").font = SF
            ws0.cell(row=row, column=3, value=f"업종: {cc}").font = DF
            if cl.get("crno"):
                ws0.cell(row=row, column=5, value=f"법인번호: {cl['crno']}").font = Font(size=9, color="999999", name="Malgun Gothic")
            row += 1

            # 재무정보
            if fin:
                ws0.cell(row=row, column=1, value="재무현황").font = Font(bold=True, size=10, name="Malgun Gothic")
                ws0.cell(row=row, column=2, value=f"({fin['기준연도']}년 {fin['재무제표구분']})").font = Font(size=9, color="888888", name="Malgun Gothic")
                row += 1

                fin_items = [
                    ("매출액", fin["매출액"]),
                    ("영업이익", fin["영업이익"]),
                    ("당기순이익", fin["당기순이익"]),
                    ("자산총계", fin["자산총계"]),
                    ("부채총계", fin["부채총계"]),
                    ("자본총계", fin["자본총계"]),
                ]
                fin_headers = [f[0] for f in fin_items]
                fin_values = [fmt_money(f[1]) + "만원" if f[1] != 0 else "-" for f in fin_items]

                for c, h in enumerate(fin_headers, 1):
                    cell = ws0.cell(row=row, column=c, value=h)
                    cell.font = Font(bold=True, size=9, color="666666", name="Malgun Gothic")
                    cell.border = BD
                row += 1
                for c, v in enumerate(fin_values, 1):
                    cell = ws0.cell(row=row, column=c, value=v)
                    cell.font = DF
                    cell.border = BD

                # 부채비율 계산
                if fin["자본총계"] > 0:
                    debt_ratio = round(fin["부채총계"] / fin["자본총계"] * 100, 1)
                    ws0.cell(row=row, column=7, value=f"부채비율: {debt_ratio}%").font = Font(
                        bold=True, size=10, name="Malgun Gothic",
                        color="C00000" if debt_ratio > 200 else "006600"
                    )
                row += 1
            else:
                ws0.cell(row=row, column=1, value="재무정보: 법인등록번호 입력 시 자동 조회됩니다").font = Font(size=10, color="999999", italic=True, name="Malgun Gothic")
                row += 1

            # 매칭 발주
            ws0.cell(row=row, column=1, value=f"매칭 가능 재량계약: {len(mc)}건").font = Font(bold=True, size=10, name="Malgun Gothic")
            row += 1

            if mc:
                mh = ["사업명", "발주기관", "예산(만원)", "업종", "지역", "발주시기", "매칭점수"]
                for c, h in enumerate(mh, 1):
                    cell = ws0.cell(row=row, column=c, value=h)
                    cell.fill, cell.font, cell.border = PatternFill("solid", fgColor="D6E4F0"), Font(bold=True, size=9, name="Malgun Gothic"), BD
                row += 1
                for m in mc[:15]:  # 최대 15건
                    for c, h in enumerate(mh, 1):
                        cell = ws0.cell(row=row, column=c, value=m.get(h, ""))
                        cell.font, cell.border = DF, BD
                    ws0.cell(row=row, column=3).number_format = "#,##0"
                    row += 1
                if len(mc) > 15:
                    ws0.cell(row=row, column=1, value=f"  ... 외 {len(mc)-15}건 (전체목록은 '고객매칭' 시트 참조)").font = Font(size=9, color="888888", italic=True, name="Malgun Gothic")
                    row += 1

            row += 2  # 고객사 간 간격

    else:
        ws0.cell(row=start_row, column=1, value="config.txt/clients.csv에 고객사를 입력하면 고객사별 수주역량 진단이 여기에 표시됩니다.").font = Font(size=11, color="888888", italic=True, name="Malgun Gothic")

    aw(ws0)

    # ═══════════════════════════════════════════
    # 시트2: 재량계약 타겟
    # ═══════════════════════════════════════════
    ws1 = wb.create_sheet("재량계약 타겟")
    ws1.sheet_properties.tabColor = "70AD47"

    disc_sorted = sorted(disc_data, key=lambda x: x["예산액"], reverse=True)
    c1 = ["사업명", "발주기관", "예산(만원)", "업종", "재량여부", "지역", "발주시기", "담당자", "연락처"]
    for c, h in enumerate(c1, 1):
        ws1.cell(row=1, column=c, value=h)
    hdr(ws1, 1, len(c1))

    for r, d in enumerate(disc_sorted, 2):
        for c, h in enumerate(c1, 1):
            cell = ws1.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = DF, BD
            if r % 2 == 0:
                cell.fill = AF
        ws1.cell(row=r, column=3).number_format = "#,##0"
        if d["재량여부"] == "소액수의":
            ws1.cell(row=r, column=5).fill = GF

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(c1))}1"
    ws1.freeze_panes = "A2"
    aw(ws1)

    # ═══════════════════════════════════════════
    # 시트3: 전체 발주계획
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("전체 발주계획")
    ws2.sheet_properties.tabColor = "4472C4"

    c2 = ["번호", "사업명", "발주기관", "수요기관", "예산(만원)", "예산(억)",
          "발주시기", "계약방법", "업종", "재량여부", "지역", "담당자", "연락처"]
    for c, h in enumerate(c2, 1):
        ws2.cell(row=1, column=c, value=h)
    hdr(ws2, 1, len(c2))

    for r, d in enumerate(all_data, 2):
        for c, h in enumerate(c2, 1):
            cell = ws2.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = DF, BD
            if r % 2 == 0:
                cell.fill = AF
        ws2.cell(row=r, column=5).number_format = "#,##0"
        ws2.cell(row=r, column=6).number_format = "#,##0.00"
        if ws2.cell(row=r, column=10).value in ("소액수의", "재량계약"):
            ws2.cell(row=r, column=10).fill = GF

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(c2))}1"
    ws2.freeze_panes = "A2"
    aw(ws2)

    # ═══════════════════════════════════════════
    # 시트4: 통계
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("통계")
    ws3.sheet_properties.tabColor = "7030A0"

    ws3["A1"] = f"발주계획 분석 ({config.get('YEAR', '')})"
    ws3["A1"].font = TF
    ws3.merge_cells("A1:E1")
    ws3["A2"] = f"전체 {len(all_data)}건 | 재량계약 {len(disc_data)}건 | {datetime.now():%Y-%m-%d %H:%M}"
    ws3["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")

    missing_owner = sum(1 for d in all_data if not d.get("담당자"))
    missing_tel = sum(1 for d in all_data if not d.get("연락처"))
    owner_pct = round((missing_owner / len(all_data) * 100), 1) if all_data else 0
    tel_pct = round((missing_tel / len(all_data) * 100), 1) if all_data else 0
    ws3["A3"] = f"연락정보 누락: 담당자 {missing_owner}건({owner_pct}%), 연락처 {missing_tel}건({tel_pct}%)"
    ws3["A3"].font = Font(size=9, color="666666", name="Malgun Gothic")

    # 업종별
    cat_st = {}
    for d in all_data:
        ct = d["업종"]
        if ct not in cat_st:
            cat_st[ct] = {"n": 0, "amt": 0, "disc": 0}
        cat_st[ct]["n"] += 1
        cat_st[ct]["amt"] += d["예산액"]
        if d["재량여부"] in ("소액수의", "재량계약"):
            cat_st[ct]["disc"] += 1

    row = 5
    sc = ["업종", "건수", "금액(억)", "재량건수", "재량비율(%)"]
    for c, h in enumerate(sc, 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, len(sc))

    for ct, st in sorted(cat_st.items(), key=lambda x: x[1]["n"], reverse=True):
        row += 1
        ws3.cell(row=row, column=1, value=ct).font = DF
        ws3.cell(row=row, column=2, value=st["n"]).font = DF
        ws3.cell(row=row, column=3, value=round(st["amt"] / 1_0000_0000, 1)).font = DF
        ws3.cell(row=row, column=4, value=st["disc"]).font = DF
        pct = round(st["disc"] / st["n"] * 100, 1) if st["n"] > 0 else 0
        ws3.cell(row=row, column=5, value=pct).font = DF
        for c in range(1, 6):
            ws3.cell(row=row, column=c).border = BD

    # 지역별
    row += 2
    ws3.cell(row=row, column=1, value="[지역별]").font = SF
    row += 1
    for c, h in enumerate(["지역", "건수", "재량건수"], 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, 3)

    rgn = {}
    for d in all_data:
        r2 = d["지역"] or "미분류"
        if r2 not in rgn:
            rgn[r2] = {"n": 0, "d": 0}
        rgn[r2]["n"] += 1
        if d["재량여부"] in ("소액수의", "재량계약"):
            rgn[r2]["d"] += 1
    for rg, st in sorted(rgn.items(), key=lambda x: x[1]["n"], reverse=True)[:15]:
        row += 1
        ws3.cell(row=row, column=1, value=rg).font = DF
        ws3.cell(row=row, column=2, value=st["n"]).font = DF
        ws3.cell(row=row, column=3, value=st["d"]).font = DF

    aw(ws3)

    # ═══════════════════════════════════════════
    # 시트5: 고객매칭 (전체)
    # ═══════════════════════════════════════════
    if matched:
        ws4 = wb.create_sheet("고객매칭")
        ws4.sheet_properties.tabColor = "ED7D31"
        c4 = ["매칭고객사", "고객업종", "매칭점수", "매칭근거", "사업명", "발주기관", "예산(만원)", "재량여부", "지역", "담당자", "연락처"]
        for c, h in enumerate(c4, 1):
            ws4.cell(row=1, column=c, value=h)
        hdr(ws4, 1, len(c4))
        for r, d in enumerate(matched, 2):
            for c, h in enumerate(c4, 1):
                cell = ws4.cell(row=r, column=c, value=d.get(h, ""))
                cell.font, cell.border = DF, BD
            ws4.cell(row=r, column=7).number_format = "#,##0"
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(c4))}1"
        ws4.freeze_panes = "A2"
        aw(ws4)

    # ═══════════════════════════════════════════
    # 시트6: 고객사 재무현황
    # ═══════════════════════════════════════════
    if financials:
        ws5 = wb.create_sheet("고객 재무현황")
        ws5.sheet_properties.tabColor = "C00000"
        c5 = ["고객사", "업종", "기준연도", "매출액(만원)", "영업이익(만원)", "당기순이익(만원)",
              "자산총계(만원)", "부채총계(만원)", "자본총계(만원)", "부채비율(%)"]
        for c, h in enumerate(c5, 1):
            ws5.cell(row=1, column=c, value=h)
        hdr(ws5, 1, len(c5))

        r = 2
        for cn, fin in financials.items():
            cl = next((c for c in clients if c["name"] == cn), {})
            ws5.cell(row=r, column=1, value=cn).font = DF
            ws5.cell(row=r, column=2, value=cl.get("category", "")).font = DF
            ws5.cell(row=r, column=3, value=fin["기준연도"]).font = DF
            ws5.cell(row=r, column=4, value=round(fin["매출액"] / 10000) if fin["매출액"] else 0).font = DF
            ws5.cell(row=r, column=5, value=round(fin["영업이익"] / 10000) if fin["영업이익"] else 0).font = DF
            ws5.cell(row=r, column=6, value=round(fin["당기순이익"] / 10000) if fin["당기순이익"] else 0).font = DF
            ws5.cell(row=r, column=7, value=round(fin["자산총계"] / 10000) if fin["자산총계"] else 0).font = DF
            ws5.cell(row=r, column=8, value=round(fin["부채총계"] / 10000) if fin["부채총계"] else 0).font = DF
            ws5.cell(row=r, column=9, value=round(fin["자본총계"] / 10000) if fin["자본총계"] else 0).font = DF
            dr = round(fin["부채총계"] / fin["자본총계"] * 100, 1) if fin["자본총계"] > 0 else 0
            cell = ws5.cell(row=r, column=10, value=dr)
            cell.font = Font(bold=True, size=10, name="Malgun Gothic", color="C00000" if dr > 200 else "006600")
            for c in range(1, 11):
                ws5.cell(row=r, column=c).border = BD
                if c >= 4:
                    ws5.cell(row=r, column=c).number_format = "#,##0"
            r += 1
        aw(ws5)

    # ═══════════════════════════════════════════
    # 시트7: 주간증감
    # ═══════════════════════════════════════════
    ws6 = wb.create_sheet("주간증감")
    ws6.sheet_properties.tabColor = "2F5597"
    ws6["A1"] = "전회차 대비 증감"
    ws6["A1"].font = TF
    ws6.merge_cells("A1:E1")

    if delta_info and delta_info.get("previous"):
        prev = delta_info["previous"]
        curr = delta_info["current"]
        ws6["A2"] = f"현재: {curr['run_ts']} | 이전: {prev['run_ts']}"
        ws6["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")
        headers = ["지표", "현재", "이전", "증감", "증감률(%)"]
        for c, h in enumerate(headers, 1):
            ws6.cell(row=4, column=c, value=h)
        hdr(ws6, 4, len(headers))

        metrics = [
            ("전체 건수", curr["total_count"], prev["total_count"]),
            ("재량 건수", curr["disc_count"], prev["disc_count"]),
            ("공사 건수", curr["const_count"], prev["const_count"]),
            ("용역 건수", curr["serv_count"], prev["serv_count"]),
        ]
        row = 5
        for name, c_val, p_val in metrics:
            delta = c_val - p_val
            delta_pct = round((delta / p_val) * 100, 1) if p_val else 0
            vals = [name, c_val, p_val, delta, delta_pct]
            for col, val in enumerate(vals, 1):
                cell = ws6.cell(row=row, column=col, value=val)
                cell.font = DF
                cell.border = BD
            row += 1
    else:
        ws6["A2"] = "이전 회차 데이터가 없어 증감 계산을 생략했습니다."
        ws6["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")

    aw(ws6)

    # 저장
    fname = f"G2B_Report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    fpath = OUTPUT_DIR / fname
    wb.save(str(fpath))
    try:
        shutil.copyfile(fpath, LATEST_XLSX_FILE)
    except Exception as e:
        log.warning(f"  latest.xlsx 갱신 실패: {e}")
    return fpath


def write_csv_file(path, rows, fieldnames):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def export_data_files(all_data, disc_data, matched, run_stamp, summary_payload):
    """CSV/JSON 동시 내보내기."""
    exports = {}

    all_json = OUTPUT_DIR / f"G2B_All_{run_stamp}.json"
    disc_json = OUTPUT_DIR / f"G2B_Disc_{run_stamp}.json"
    match_json = OUTPUT_DIR / f"G2B_Matched_{run_stamp}.json"
    summary_json = OUTPUT_DIR / f"G2B_Summary_{run_stamp}.json"

    all_json.write_text(json.dumps(all_data, ensure_ascii=False, indent=2), encoding="utf-8")
    disc_json.write_text(json.dumps(disc_data, ensure_ascii=False, indent=2), encoding="utf-8")
    match_json.write_text(json.dumps(matched, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_json.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    all_csv = OUTPUT_DIR / f"G2B_All_{run_stamp}.csv"
    disc_csv = OUTPUT_DIR / f"G2B_Disc_{run_stamp}.csv"
    match_csv = OUTPUT_DIR / f"G2B_Matched_{run_stamp}.csv"

    def export_or_empty_csv(path, rows):
        if rows:
            fields = list(rows[0].keys())
            write_csv_file(path, rows, fields)
        else:
            # Keep output contract stable even when there is no row.
            path.write_text("", encoding="utf-8-sig")

    export_or_empty_csv(all_csv, all_data)
    export_or_empty_csv(disc_csv, disc_data)
    export_or_empty_csv(match_csv, matched)
    exports["all_csv"] = str(all_csv)
    exports["disc_csv"] = str(disc_csv)
    exports["matched_csv"] = str(match_csv)

    for src, latest_name in [
        (all_json, "latest_all.json"),
        (disc_json, "latest_disc.json"),
        (match_json, "latest_matched.json"),
        (summary_json, "latest_summary.json"),
        (all_csv, "latest_all.csv"),
        (disc_csv, "latest_disc.csv"),
        (match_csv, "latest_matched.csv"),
    ]:
        try:
            shutil.copyfile(src, OUTPUT_DIR / latest_name)
        except Exception:
            pass

    exports.update(
        {
            "all_json": str(all_json),
            "disc_json": str(disc_json),
            "matched_json": str(match_json),
            "summary_json": str(summary_json),
        }
    )
    return exports


def init_history_db():
    conn = sqlite3.connect(SQLITE_DB_FILE)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS run_history (
            run_id TEXT PRIMARY KEY,
            run_ts TEXT NOT NULL,
            year INTEGER NOT NULL,
            total_count INTEGER NOT NULL,
            disc_count INTEGER NOT NULL,
            const_count INTEGER NOT NULL,
            serv_count INTEGER NOT NULL,
            quality_level TEXT,
            quality_detail TEXT,
            key_source TEXT,
            profile_info TEXT,
            output_xlsx TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS plan_items (
            run_id TEXT NOT NULL,
            number TEXT,
            biz_name TEXT,
            order_org TEXT,
            category TEXT,
            contract_type TEXT,
            budget_amt INTEGER,
            region TEXT,
            contact_name TEXT,
            contact_tel TEXT,
            match_client TEXT,
            match_score INTEGER
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_run_history_year_ts ON run_history(year, run_ts DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_plan_items_run_id ON plan_items(run_id)")
    conn.commit()
    return conn


def get_previous_run_summary(conn, year):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT run_id, run_ts, total_count, disc_count, const_count, serv_count
        FROM run_history
        WHERE year = ?
        ORDER BY run_ts DESC
        LIMIT 1
        """,
        (int(year),),
    )
    row = cur.fetchone()
    if not row:
        return None
    return {
        "run_id": row[0],
        "run_ts": row[1],
        "total_count": row[2],
        "disc_count": row[3],
        "const_count": row[4],
        "serv_count": row[5],
    }


def save_run_history(conn, run_summary, all_data, matched):
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO run_history (
            run_id, run_ts, year, total_count, disc_count, const_count, serv_count,
            quality_level, quality_detail, key_source, profile_info, output_xlsx
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            run_summary["run_id"],
            run_summary["run_ts"],
            run_summary["year"],
            run_summary["total_count"],
            run_summary["disc_count"],
            run_summary["const_count"],
            run_summary["serv_count"],
            run_summary["quality_level"],
            run_summary["quality_detail"],
            run_summary["key_source"],
            run_summary["profile_info"],
            run_summary["output_xlsx"],
        ),
    )

    # 같은 공고의 최고 매칭점수만 반영
    best_match = {}
    for m in matched:
        key = (m.get("번호", ""), m.get("사업명", ""), m.get("발주기관", ""))
        prev = best_match.get(key)
        if not prev or m.get("매칭점수", 0) > prev.get("매칭점수", 0):
            best_match[key] = m

    rows = []
    for d in all_data:
        key = (d.get("번호", ""), d.get("사업명", ""), d.get("발주기관", ""))
        mm = best_match.get(key, {})
        rows.append(
            (
                run_summary["run_id"],
                d.get("번호", ""),
                d.get("사업명", ""),
                d.get("발주기관", ""),
                d.get("업종", ""),
                d.get("재량여부", ""),
                int(d.get("예산액", 0) or 0),
                d.get("지역", ""),
                d.get("담당자", ""),
                d.get("연락처", ""),
                mm.get("매칭고객사", ""),
                int(mm.get("매칭점수", 0) or 0),
            )
        )
    cur.executemany(
        """
        INSERT INTO plan_items (
            run_id, number, biz_name, order_org, category, contract_type, budget_amt,
            region, contact_name, contact_tel, match_client, match_score
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()


def build_delta_info(current_summary, previous_summary):
    if not previous_summary:
        return {"current": current_summary, "previous": None}
    return {"current": current_summary, "previous": previous_summary}


def send_webhook(webhook_url, payload):
    if not webhook_url:
        return
    try:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = Request(
            webhook_url,
            data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urlopen(req, timeout=8):
            pass
    except Exception as e:
        log.warning(f"  웹훅 전송 실패: {e}")


def validate_configuration(config, args):
    errors = []
    warnings = []

    enc_key, dec_key, key_sources = resolve_api_keys(config)
    year = resolve_target_year(config if args.year is None else {**config, "YEAR": str(args.year)})
    clients = parse_clients(config)

    if (not enc_key or "여기에" in enc_key) and (not dec_key or "여기에" in dec_key):
        errors.append("API 키 누락: config.txt 또는 환경변수 ENCODING_KEY/DECODING_KEY 필요")
    if args.max_pages <= 0:
        errors.append("--max-pages는 1 이상이어야 합니다.")
    if args.rows <= 0:
        errors.append("--rows는 1 이상이어야 합니다.")
    if not clients:
        warnings.append("고객사가 비어 있습니다. CLIENTS 또는 clients.csv를 확인하세요.")
    if CLIENTS_CSV_FILE.exists() and not load_clients_csv():
        warnings.append("clients.csv 파일이 있으나 로드된 고객사가 없습니다(헤더/컬럼명 확인).")

    return {
        "errors": errors,
        "warnings": warnings,
        "key_sources": key_sources,
        "year": year,
        "client_count": len(clients),
    }


# =============================================================
#  메인
# =============================================================

def main(
    non_interactive=False,
    year_override=None,
    max_pages=30,
    rows=999,
    check_config=False,
    refresh_profile=False,
    webhook_url_override="",
):
    interactive = not non_interactive

    print()
    print("=" * 56)
    print("  건설업 수주역량 자동진단 시스템 v3.0")
    print("  행정사사무소 하랑 · 서울건설정보")
    print("=" * 56)
    print()

    # ── 설정 ──
    config = load_config()
    if config is None:
        print("  config.txt를 생성했습니다.")
        if not interactive:
            print("  [X] 비대화 모드에서는 최초 설정을 진행할 수 없습니다.")
            print(f"      config.txt를 먼저 작성하세요: {CONFIG_FILE}")
            return 1
        print("  인코딩/디코딩 키를 입력하세요.")
        print()
        if sys.platform == "win32":
            os.system(f'notepad "{CONFIG_FILE}"')
        wait_for_enter("  키 입력 후 Enter...", interactive=True)
        config = load_config()

    ensure_clients_csv_template()

    # ── 운영 파라미터 ──
    if year_override is None:
        target_year = resolve_target_year(config)
    else:
        target_year = resolve_target_year({**config, "YEAR": str(year_override)})

    retention_days = int(config.get("LOG_RETENTION_DAYS", "30") or 30)
    cleanup_old_logs(retention_days=retention_days)

    webhook_url = (
        webhook_url_override
        or (os.environ.get("WEBHOOK_URL", "") or "").strip()
        or (config.get("WEBHOOK_URL", "") or "").strip()
    )

    args_like = argparse.Namespace(year=target_year, max_pages=max_pages, rows=rows)
    validation = validate_configuration(config, args_like)
    if check_config:
        print("  [CHECK-CONFIG]")
        print(f"  - year={validation['year']}, clients={validation['client_count']}")
        print(f"  - key_source: encoding={validation['key_sources']['encoding']}, decoding={validation['key_sources']['decoding']}")
        for w in validation["warnings"]:
            print(f"  [WARN] {w}")
        for e in validation["errors"]:
            print(f"  [ERR] {e}")
        rc = 1 if validation["errors"] else 0
        if rc:
            send_webhook(
                webhook_url,
                {
                    "event": "g2b_run_failed",
                    "reason": "config_check_error",
                    "errors": validation["errors"],
                    "ts": datetime.now().isoformat(timespec="seconds"),
                    "year": target_year,
                },
            )
        return rc

    if validation["errors"]:
        print("  [X] 설정 오류")
        for e in validation["errors"]:
            print(f"  - {e}")
        send_webhook(
            webhook_url,
            {
                "event": "g2b_run_failed",
                "reason": "config_validation_error",
                "errors": validation["errors"],
                "ts": datetime.now().isoformat(timespec="seconds"),
                "year": target_year,
            },
        )
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    enc_key, dec_key, key_sources = resolve_api_keys(config)
    clients = parse_clients(config)

    log.info("  설정 OK")
    log.info(f"  키 소스: encoding={key_sources['encoding']}, decoding={key_sources['decoding']}")
    log.info(f"  조회연도: {target_year}")
    log.info(f"  수집옵션: max_pages={max_pages}, rows={rows}")

    if clients:
        log.info(f"  고객사 {len(clients)}곳: {', '.join(c['name'] for c in clients)}")

    # ── 런타임 캐시 (성공 endpoint/profile 조합) ──
    profile_cache = load_g2b_profile_cache()
    year_key = str(target_year)
    year_cache = {} if refresh_profile else profile_cache.get(year_key, {})
    if not isinstance(year_cache, dict):
        year_cache = {}
    if refresh_profile:
        log.info(f"  프로파일 캐시 무시: --refresh-profile")
    elif year_cache:
        log.info(f"  프로파일 캐시 발견: year={year_key}")

    # ── 1. 나라장터 발주계획 수집 ──
    log.info("")
    log.info("  [1/6] 나라장터 발주계획 수집")
    log.info("")

    raw = []
    endpoint_counts = {}
    endpoint_meta = {}
    for ep in ["const", "serv"]:
        items, meta = g2b_fetch_all(
            enc_key,
            dec_key,
            ep,
            year=target_year,
            max_pages=max_pages,
            preferred=year_cache.get(ep),
            rows=rows,
        )
        endpoint_meta[ep] = meta
        endpoint_counts[ep] = len(items)
        raw.extend(items)
        if meta.get("selected_endpoint") and meta.get("selected_profile"):
            year_cache[ep] = {
                "endpoint": meta["selected_endpoint"],
                "profile": meta["selected_profile"],
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            }
        if meta.get("page_error"):
            log.warning(f"  [{G2B_LABEL[ep]}] 페이지 수집 일부 중단: {meta['page_error']}")
        time.sleep(1)

    profile_cache[year_key] = year_cache
    save_g2b_profile_cache(profile_cache)

    if not raw:
        print()
        print("  [X] 나라장터 데이터 수집 실패")
        print("  - 공공데이터포털 > 마이페이지 > 인증키 확인")
        print("  - '발주계획현황서비스' 활용신청 여부 확인")
        print("  - 필수 조회 파라미터(예: inqryDiv/orderBgnYm/orderEndYm) 누락 여부 확인")
        print("  - 키 포함 호출에서 HTTP 404 반복 시 엔드포인트 변경 가능성도 점검")
        send_webhook(
            webhook_url,
            {
                "event": "g2b_run_failed",
                "reason": "no_data",
                "ts": datetime.now().isoformat(timespec="seconds"),
                "year": target_year,
                "endpoint_counts": endpoint_counts,
            },
        )
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    for ep, n in endpoint_counts.items():
        log.info(f"  [{G2B_LABEL[ep]}] 수집 건수: {n}건")

    success_eps = [ep for ep, n in endpoint_counts.items() if n > 0]
    if len(success_eps) == len(endpoint_counts):
        quality_level = "FULL"
    elif success_eps:
        quality_level = "PARTIAL"
    else:
        quality_level = "FAIL"
    quality_detail = " | ".join(f"{G2B_LABEL[ep]} {endpoint_counts[ep]}건" for ep in ["const", "serv"])
    data_quality = {"level": quality_level, "detail": quality_detail}
    log.info(f"  DATA_QUALITY={quality_level} ({quality_detail})")

    # ── 2. 데이터 처리 ──
    log.info("")
    log.info("  [2/6] 데이터 처리")

    all_data = process_g2b(raw, config)
    disc_data = [d for d in all_data if d["재량여부"] in ("소액수의", "재량계약")]

    cat_n = {}
    for d in all_data:
        cat_n[d["업종"]] = cat_n.get(d["업종"], 0) + 1

    log.info(f"  전체: {len(all_data)}건 | 재량계약: {len(disc_data)}건")
    for ct, n in sorted(cat_n.items(), key=lambda x: x[1], reverse=True):
        log.info(f"    {ct}: {n}건")

    # ── 3. 고객사 재무정보 + 매칭 ──
    log.info("")
    log.info("  [3/6] 고객사 분석")

    financials = {}
    matched = []

    if clients:
        matched = match_clients(disc_data, clients)
        log.info(f"  발주 매칭: {len(matched)}건")

        # 법인등록번호가 있는 고객사만 재무조회
        crno_clients = [c for c in clients if c.get("crno")]
        if crno_clients:
            log.info(f"  재무정보 조회: {len(crno_clients)}사")
            for cl in crno_clients:
                log.info(f"    {cl['name']}...")
                fin = get_client_financials(enc_key, dec_key, cl)
                if fin:
                    financials[cl["name"]] = fin
                    log.info(f"    -> {fin['기준연도']}년 매출 {fmt_eok_simple(fin['매출액'])}억")
                else:
                    log.info(f"    -> 조회 실패 (법인번호 확인 필요)")
                time.sleep(0.5)
        else:
            log.info("  법인등록번호가 있는 고객사 없음 - 재무조회 건너뜀")
    else:
        log.info("  등록된 고객사 없음")

    # ── 4. 저장소(SQLite) 증감 계산 ──
    log.info("")
    log.info("  [4/6] 이력 분석")
    conn = init_history_db()
    try:
        previous_summary = get_previous_run_summary(conn, target_year)
    finally:
        conn.close()

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_ts = datetime.now().isoformat(timespec="seconds")
    profile_info = " | ".join(
        f"{G2B_LABEL[ep]}:{endpoint_meta.get(ep, {}).get('selected_endpoint', '-')}/{endpoint_meta.get(ep, {}).get('selected_profile', '-')}"
        for ep in ["const", "serv"]
    )
    key_source_text = f"encoding={key_sources['encoding']},decoding={key_sources['decoding']}"
    current_summary = {
        "run_id": run_id,
        "run_ts": run_ts,
        "year": int(target_year),
        "total_count": len(all_data),
        "disc_count": len(disc_data),
        "const_count": endpoint_counts.get("const", 0),
        "serv_count": endpoint_counts.get("serv", 0),
        "quality_level": quality_level,
        "quality_detail": quality_detail,
        "key_source": key_source_text,
        "profile_info": profile_info,
        "output_xlsx": "",
    }
    delta_info = build_delta_info(current_summary, previous_summary)

    # ── 5. 리포트 생성 ──
    log.info("")
    log.info("  [5/6] 리포트 생성")

    if not ensure_openpyxl():
        send_webhook(
            webhook_url,
            {
                "event": "g2b_run_failed",
                "reason": "dependency_error",
                "dependency": "openpyxl",
                "ts": datetime.now().isoformat(timespec="seconds"),
                "year": target_year,
            },
        )
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    runtime_meta = {
        "run_mode": "non-interactive" if non_interactive else "interactive",
        "key_source": key_source_text,
        "profile_info": profile_info,
        "target_year": target_year,
    }
    fpath = create_report(
        all_data,
        disc_data,
        matched,
        clients,
        financials,
        config,
        data_quality=data_quality,
        runtime_meta=runtime_meta,
        delta_info=delta_info,
    )
    current_summary["output_xlsx"] = str(fpath)

    # ── 6. 산출물 내보내기 + DB 적재 ──
    log.info("")
    log.info("  [6/6] 내보내기/이력적재")
    summary_payload = {
        "run_id": run_id,
        "run_ts": run_ts,
        "year": target_year,
        "quality": data_quality,
        "endpoint_counts": endpoint_counts,
        "profile_info": profile_info,
        "key_source": key_source_text,
        "counts": {
            "all": len(all_data),
            "disc": len(disc_data),
            "matched": len(matched),
            "financials": len(financials),
        },
        "output_xlsx": str(fpath),
    }
    exports = export_data_files(all_data, disc_data, matched, run_id, summary_payload)
    log.info(f"  JSON/CSV 내보내기 완료: {len(exports)}개 파일")

    conn = init_history_db()
    try:
        save_run_history(conn, current_summary, all_data, matched)
    finally:
        conn.close()

    send_webhook(
        webhook_url,
        {
            "event": "g2b_run_succeeded",
            "ts": run_ts,
            "year": target_year,
            "quality": data_quality,
            "counts": summary_payload["counts"],
            "output_xlsx": str(fpath),
        },
    )

    # ── 완료 ──
    print()
    print("=" * 56)
    print("  [OK] 완료!")
    print(f"  파일: {fpath}")
    print(f"  전체 {len(all_data)}건 | 재량계약 {len(disc_data)}건")
    if matched:
        print(f"  고객매칭 {len(matched)}건")
    if financials:
        print(f"  재무조회 {len(financials)}사")
    print("=" * 56)
    print()

    if interactive and sys.platform == "win32":
        os.startfile(str(OUTPUT_DIR))

    wait_for_enter("  Enter로 종료...", interactive=interactive)
    return 0


def fmt_eok_simple(v):
    if v == 0:
        return "0"
    return f"{v / 1_0000_0000:,.1f}"


def parse_args():
    parser = argparse.ArgumentParser(description="G2B 발주계획 자동 수집 및 리포트 생성")
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="입력 대기/메모장 열기/결과 폴더 자동 열기를 비활성화합니다.",
    )
    parser.add_argument("--year", type=int, default=None, help="조회 연도 오버라이드 (예: 2026)")
    parser.add_argument("--max-pages", type=int, default=30, help="엔드포인트당 최대 페이지 수")
    parser.add_argument("--rows", type=int, default=999, help="페이지당 요청 건수")
    parser.add_argument("--check-config", action="store_true", help="실행 전 설정만 검증하고 종료합니다.")
    parser.add_argument("--refresh-profile", action="store_true", help="endpoint/profile 캐시를 무시하고 재탐색합니다.")
    parser.add_argument("--webhook-url", default="", help="실패/성공 알림 웹훅 URL 오버라이드")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        raise SystemExit(
            main(
                non_interactive=args.non_interactive,
                year_override=args.year,
                max_pages=args.max_pages,
                rows=args.rows,
                check_config=args.check_config,
                refresh_profile=args.refresh_profile,
                webhook_url_override=args.webhook_url,
            )
        )
    except KeyboardInterrupt:
        print("\n  중단")
        wait_for_enter("\n  Enter로 종료...", interactive=not args.non_interactive)
        raise SystemExit(130)
    except Exception as e:
        log.error(f"  오류: {e}")
        import traceback
        traceback.print_exc()
        try:
            cfg = load_config() or {}
            webhook_url = args.webhook_url or (os.environ.get("WEBHOOK_URL", "") or "").strip() or cfg.get("WEBHOOK_URL", "")
            send_webhook(
                webhook_url,
                {
                    "event": "g2b_run_failed",
                    "reason": "unhandled_exception",
                    "error": str(e),
                    "ts": datetime.now().isoformat(timespec="seconds"),
                },
            )
        except Exception:
            pass
        wait_for_enter("\n  Enter로 종료...", interactive=not args.non_interactive)
        raise SystemExit(1)
