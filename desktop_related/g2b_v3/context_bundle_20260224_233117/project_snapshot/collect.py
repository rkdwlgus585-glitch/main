#!/usr/bin/env python3
"""
=============================================================
  건설업 수주역량 자동진단 시스템 v3.0
  행정사사무소 하랑 · 서울건설정보
=============================================================

  [1] 나라장터 발주계획 자동수집 + 업종분류 + 재량계약 필터
  [2] 금융위 기업재무정보 → 고객사 재무 자동조회
  [3] 고객사 × 발주계획 매칭
  [4] 수주역량 진단 리포트 (엑셀)

  ★ config.txt에 키만 넣으면 나머지는 전자동
  ★ run.bat 더블클릭으로 실행
=============================================================
"""

import os, sys, json, time, logging, argparse
from datetime import datetime
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.parse import urlencode, quote
from urllib.error import URLError, HTTPError

# ── 경로 ──
BASE_DIR = Path(__file__).parent.resolve()
CONFIG_FILE = BASE_DIR / "config.txt"
OUTPUT_DIR = BASE_DIR / "result"
LOG_DIR = BASE_DIR / "logs"

LOG_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(LOG_DIR / f"log_{datetime.now():%Y%m%d}.txt"), encoding="utf-8"),
    ]
)
log = logging.getLogger()


# =============================================================
#  설정
# =============================================================

DEFAULT_CONFIG = """# ============================================
#  건설업 수주역량 자동진단 설정파일
#  행정사사무소 하랑 · 서울건설정보
# ============================================
#
# ★ 공공데이터포털 마이페이지 > 인증키에서
#   Encoding 키와 Decoding 키 둘 다 넣으세요
#   (프로그램이 자동으로 맞는 키를 찾습니다)
#

ENCODING_KEY=여기에_인코딩키_붙여넣기
DECODING_KEY=여기에_디코딩키_붙여넣기

# ── 아래는 수정 안해도 됩니다 ──
YEAR={year}
DISC_LIMIT=20000
SMALL_LIMIT=5000

# ── 인큐베이팅 고객사 ──
# 형식: 회사명:업종:법인등록번호
# 법인등록번호가 있으면 재무정보를 자동 조회합니다
# 법인등록번호가 없으면 :만 쓰거나 비워두세요
#
# 예시:
# CLIENTS=하늘건설:일반건설:1101111234567,동방전기:전기공사:1301110001234,서울ICT:정보통신:
CLIENTS=
"""


def load_config():
    if not CONFIG_FILE.exists():
        txt = DEFAULT_CONFIG.format(year=datetime.now().year)
        CONFIG_FILE.write_text(txt, encoding="utf-8-sig")
        return None

    config = {}
    for line in CONFIG_FILE.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, v = line.split("=", 1)
            config[k.strip()] = v.strip()
    return config


def parse_clients(config):
    """고객사 파싱 → [{name, category, crno}, ...]"""
    raw = config.get("CLIENTS", "")
    if not raw:
        return []
    clients = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        segs = part.split(":")
        name = segs[0].strip() if len(segs) > 0 else ""
        cat = segs[1].strip() if len(segs) > 1 else ""
        crno = segs[2].strip() if len(segs) > 2 else ""
        if name:
            clients.append({"name": name, "category": cat, "crno": crno})
    return clients


def wait_for_enter(message, interactive):
    """Interactive 모드에서만 입력 대기."""
    if not interactive:
        return
    try:
        input(message)
    except EOFError:
        log.warning("  비대화 환경으로 감지되어 입력 대기를 건너뜁니다.")


# =============================================================
#  공통 API 호출
# =============================================================

def try_api_call(url_template, encoding_key, decoding_key, extra_params=None, call_label="API"):
    """
    인코딩키 → 디코딩키 순서로 시도.
    조달청은 인코딩키, 금융위는 디코딩키를 주로 씀.
    """
    failures = []
    for key_label, key_val in [("encoding", encoding_key), ("decoding", decoding_key)]:
        if not key_val or "여기에" in key_val:
            continue
        try:
            params = {"serviceKey": key_val}
            if extra_params:
                params.update(extra_params)

            # 조달청 API는 serviceKey를 직접 URL에 넣어야 함 (이미 인코딩된 키)
            if key_label == "encoding":
                # 인코딩키는 이미 URL인코딩 되어있으므로 직접 붙임
                param_str = f"serviceKey={key_val}"
                other_params = {k: v for k, v in params.items() if k != "serviceKey"}
                if other_params:
                    param_str += "&" + urlencode(other_params)
                full_url = f"{url_template}?{param_str}"
            else:
                # 디코딩키는 urlencode가 자동 인코딩
                full_url = f"{url_template}?{urlencode(params)}"

            req = Request(full_url, headers={"Accept": "application/json"})
            with urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8")

            # XML 에러 응답 체크
            if raw.strip().startswith("<"):
                if "SERVICE_KEY_IS_NOT_REGISTERED_ERROR" in raw:
                    continue
                if "UNREGISTERED_SERVICE_ERROR" in raw:
                    continue
                if "<resultCode>00</resultCode>" not in raw and "<resultCode>0</resultCode>" not in raw:
                    continue

            data = json.loads(raw)
            return data, key_label

        except (json.JSONDecodeError, UnicodeDecodeError):
            failures.append(f"{key_label}:invalid-response")
            continue
        except HTTPError as e:
            if e.code in (401, 403, 404, 500, 502, 503):
                log.warning(f"  HTTP {e.code} ({key_label}키) - 다음 키 시도")
                failures.append(f"{key_label}:HTTP{e.code}")
                continue
            failures.append(f"{key_label}:HTTP{e.code}")
            raise
        except Exception as e:
            failures.append(f"{key_label}:{type(e).__name__}")
            continue

    if failures:
        uniq = []
        for f in failures:
            if f not in uniq:
                uniq.append(f)
        log.error(f"  {call_label} 호출 실패: {', '.join(uniq)}")
        if all("HTTP404" in f for f in uniq):
            log.error("  키 포함 호출에서 404가 반복됩니다. 활용신청 승인/키 재발급 상태를 확인하세요.")

    return None, None


# =============================================================
#  나라장터 발주계획 API
# =============================================================

G2B_BASE = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService"
G2B_EPS = {
    "const": "getOrderPlanSttusListCnstwk",
    "serv": "getOrderPlanSttusListServc",
}
G2B_LABEL = {"const": "공사", "serv": "용역"}


def g2b_fetch_page(enc_key, dec_key, ep_key, page=1, rows=999):
    url = f"{G2B_BASE}/{G2B_EPS[ep_key]}"
    params = {"pageNo": str(page), "numOfRows": str(rows), "type": "json"}

    data, used_key = try_api_call(
        url, enc_key, dec_key, params, call_label=f"G2B:{G2B_EPS[ep_key]}"
    )
    if not data:
        return [], 0, None

    if "response" in data:
        header = data["response"].get("header", {})
        if header.get("resultCode") != "00":
            log.error(f"  G2B 오류: {header.get('resultMsg', '?')}")
            return [], 0, used_key
        body = data["response"].get("body", {})
        items = body.get("items", [])
        if isinstance(items, dict):
            items = items.get("item", [])
        if isinstance(items, dict):
            items = [items]
        return items, int(body.get("totalCount", 0)), used_key
    return [], 0, None


def g2b_fetch_all(enc_key, dec_key, ep_key, max_pages=30):
    label = G2B_LABEL[ep_key]
    all_items = []
    page = 1
    total = 0
    working_key_type = None

    while page <= max_pages:
        log.info(f"  [{label}] {page}페이지...")
        items, total, kt = g2b_fetch_page(enc_key, dec_key, ep_key, page)

        if kt and not working_key_type:
            working_key_type = kt
            log.info(f"  [{label}] {kt} 키로 연결 성공")

        if not items:
            break
        all_items.extend(items)

        if len(all_items) >= total:
            break
        page += 1
        time.sleep(0.5)

    log.info(f"  [{label}] {len(all_items)}건 수집 (전체 {total}건)")
    return all_items


# =============================================================
#  금융위 기업 재무정보 API
# =============================================================

FSC_BASE = "https://apis.data.go.kr/1160100/service/GetFinaStatInfoService_V2"
FSC_EPS = {
    "summary": "getSummFinaStat_V2",      # 요약재무제표
    "bs": "getBs_V2",                      # 재무상태표
    "income": "getIncoStat_V2",            # 손익계산서
}


def fsc_fetch(enc_key, dec_key, endpoint, crno="", corp_name="", biz_year=""):
    """금융위 재무정보 조회"""
    url = f"{FSC_BASE}/{FSC_EPS[endpoint]}"
    params = {
        "pageNo": "1",
        "numOfRows": "20",
        "resultType": "json",
    }
    if crno:
        params["crno"] = crno
    if corp_name:
        params["fnccmpNm"] = corp_name
    if biz_year:
        params["bizYear"] = biz_year

    data, used_key = try_api_call(
        url, enc_key, dec_key, params, call_label=f"FSC:{FSC_EPS[endpoint]}"
    )
    if not data:
        return []

    try:
        items = data.get("response", {}).get("body", {}).get("items", {}).get("item", [])
        if isinstance(items, dict):
            items = [items]
        return items
    except (AttributeError, TypeError):
        return []


def get_client_financials(enc_key, dec_key, client):
    """고객사 1곳의 재무 요약 조회"""
    crno = client.get("crno", "")
    name = client.get("name", "")

    if not crno and not name:
        return None

    # 최근 2개년 시도
    current_year = datetime.now().year
    results = []

    for year in [current_year - 1, current_year - 2]:
        items = fsc_fetch(enc_key, dec_key, "summary",
                         crno=crno, corp_name=name if not crno else "",
                         biz_year=str(year))
        if items:
            results.extend(items)

    if not results:
        # 연도 없이 재시도
        items = fsc_fetch(enc_key, dec_key, "summary",
                         crno=crno, corp_name=name if not crno else "")
        if items:
            results = items

    if not results:
        return None

    # 가장 최신 데이터 선택 (개별재무제표 우선)
    best = None
    for item in results:
        dcd = item.get("fnclDcdNm", "")
        if "개별" in dcd or "별도" in dcd:
            if not best or item.get("bizYear", "") > best.get("bizYear", ""):
                best = item
    if not best and results:
        best = max(results, key=lambda x: x.get("bizYear", ""))

    if not best:
        return None

    def safe_int(v):
        try:
            return int(v) if v else 0
        except (ValueError, TypeError):
            return 0

    return {
        "기준연도": best.get("bizYear", ""),
        "재무제표구분": best.get("fnclDcdNm", ""),
        "매출액": safe_int(best.get("enpSaleAmt", 0)),
        "영업이익": safe_int(best.get("enpBzopPft", 0)),
        "당기순이익": safe_int(best.get("iclsPalClcAmt", 0)),
        "자산총계": safe_int(best.get("enpTastAmt", 0)),
        "부채총계": safe_int(best.get("enpTdbtAmt", 0)),
        "자본총계": safe_int(best.get("enpTcptlAmt", 0)),
    }


# =============================================================
#  데이터 처리
# =============================================================

CAT_RULES = {
    "전기공사": ["전기", "전력", "배전", "수변전", "조명", "전기설비", "전기공사"],
    "정보통신공사": ["정보통신", "통신", "네트워크", "CCTV", "방송", "통합배선", "ICT", "cctv"],
    "소방공사": ["소방", "방재", "소화", "경보", "스프링클러", "소방설비"],
    "기계설비공사": ["기계설비", "냉난방", "공조", "배관", "급배수"],
    "조경공사": ["조경", "녹화", "수목", "식재"],
    "건축공사": ["건축", "신축", "증축", "리모델링", "개보수", "인테리어"],
    "토목공사": ["토목", "도로", "상하수도", "하천", "교량", "포장"],
}


def classify(name):
    if not name:
        return "기타"
    u = name.upper()
    for cat, kws in CAT_RULES.items():
        for kw in kws:
            if kw.upper() in u:
                return cat
    return "기타"


def contract_type(amt, disc_limit, small_limit):
    if amt <= 0:
        return "금액미정"
    if amt <= small_limit:
        return "소액수의"
    if amt <= disc_limit:
        return "재량계약"
    return "일반경쟁"


def process_g2b(raw_items, config):
    disc = int(config.get("DISC_LIMIT", "20000")) * 10000
    small = int(config.get("SMALL_LIMIT", "5000")) * 10000

    results = []
    for item in raw_items:
        # 차세대 API: sumOrderAmt, 구버전: asignBdgtAmt
        try:
            amt = int(item.get("sumOrderAmt", 0) or item.get("asignBdgtAmt", 0) or 0)
        except (ValueError, TypeError):
            amt = 0

        # 차세대 API: bizNm, 구버전: bidNtceNm
        name = item.get("bizNm", "") or item.get("bidNtceNm", "") or item.get("prdctClsfcNoNm", "") or ""
        results.append({
            "번호": item.get("orderPlanUntyNo", ""),
            "사업명": name,
            "발주기관": item.get("orderInsttNm", ""),
            "수요기관": item.get("totlmngInsttNm", "") or item.get("dminsttNm", ""),
            "예산액": amt,
            "예산(만원)": round(amt / 10000) if amt > 0 else 0,
            "예산(억)": round(amt / 1_0000_0000, 2) if amt > 0 else 0,
            "발주시기": item.get("nticeDt", "") or item.get("orderPlanRegDt", ""),
            "계약방법": item.get("cntrctMthdNm", ""),
            "업종": classify(name),
            "재량여부": contract_type(amt, disc, small),
            "지역": item.get("cnstwkRgnNm", "") or item.get("dminsttRgnNm", "") or item.get("orderInsttRgnNm", ""),
            "담당자": item.get("ofclNm", "") or item.get("orderInsttOfclNm", ""),
            "연락처": item.get("telNo", "") or item.get("orderInsttOfclTelNo", ""),
        })
    return results


def match_clients(disc_data, clients):
    """고객사 × 재량계약 발주 매칭"""
    matched = []
    for cl in clients:
        cc = cl["category"]
        for d in disc_data:
            ic = d["업종"]
            ok = False
            if ("일반" in cc or "건축" in cc) and ("건축" in ic or "토목" in ic):
                ok = True
            elif "전기" in cc and "전기" in ic:
                ok = True
            elif "통신" in cc and "통신" in ic:
                ok = True
            elif "소방" in cc and "소방" in ic:
                ok = True
            elif "기계" in cc and "기계" in ic:
                ok = True
            elif "조경" in cc and "조경" in ic:
                ok = True
            if ok:
                matched.append({**d, "매칭고객사": cl["name"], "고객업종": cl["category"]})
    return matched


# =============================================================
#  엑셀 리포트 생성
# =============================================================

def ensure_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        log.info("  openpyxl 설치중...")
        os.system(f'"{sys.executable}" -m pip install openpyxl -q')
        try:
            import openpyxl
            return True
        except ImportError:
            return False


def create_report(all_data, disc_data, matched, clients, financials, config):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # 스타일
    HF = PatternFill("solid", fgColor="1F4E79")
    HN = Font(bold=True, color="FFFFFF", size=10, name="Malgun Gothic")
    DF = Font(size=10, name="Malgun Gothic")
    TF = Font(bold=True, size=14, name="Malgun Gothic", color="1F4E79")
    SF = Font(bold=True, size=12, name="Malgun Gothic", color="1F4E79")
    AF = PatternFill("solid", fgColor="F2F7FB")
    GF = PatternFill("solid", fgColor="E2EFDA")
    WF = PatternFill("solid", fgColor="FFF2CC")
    BD = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def hdr(ws, row, n):
        for c in range(1, n + 1):
            cl = ws.cell(row=row, column=c)
            cl.fill, cl.font, cl.border = HF, HN, BD
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def aw(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 3, 10), 42)

    def fmt_money(v):
        """원 → 만원 표시"""
        if v == 0:
            return "-"
        return f"{round(v / 10000):,}"

    def fmt_eok(v):
        """원 → 억원 표시"""
        if v == 0:
            return "-"
        return f"{v / 1_0000_0000:,.1f}"

    # ═══════════════════════════════════════════
    # 시트1: 수주역량 진단 (고객사별 요약)
    # ═══════════════════════════════════════════
    ws0 = wb.active
    ws0.title = "수주역량 진단"
    ws0.sheet_properties.tabColor = "0F4C75"

    ws0["A1"] = "건설업 수주역량 진단 리포트"
    ws0["A1"].font = TF
    ws0.merge_cells("A1:H1")
    ws0["A2"] = f"행정사사무소 하랑 | {datetime.now():%Y-%m-%d %H:%M} | {config.get('YEAR', '')}년"
    ws0["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")
    ws0.merge_cells("A2:H2")

    if clients:
        row = 4
        for cl in clients:
            cn = cl["name"]
            cc = cl["category"]
            fin = financials.get(cn)
            mc = [m for m in matched if m["매칭고객사"] == cn]

            # 고객사 헤더
            ws0.cell(row=row, column=1, value=f"■ {cn}").font = SF
            ws0.cell(row=row, column=3, value=f"업종: {cc}").font = DF
            if cl.get("crno"):
                ws0.cell(row=row, column=5, value=f"법인번호: {cl['crno']}").font = Font(size=9, color="999999", name="Malgun Gothic")
            row += 1

            # 재무정보
            if fin:
                ws0.cell(row=row, column=1, value="재무현황").font = Font(bold=True, size=10, name="Malgun Gothic")
                ws0.cell(row=row, column=2, value=f"({fin['기준연도']}년 {fin['재무제표구분']})").font = Font(size=9, color="888888", name="Malgun Gothic")
                row += 1

                fin_items = [
                    ("매출액", fin["매출액"]),
                    ("영업이익", fin["영업이익"]),
                    ("당기순이익", fin["당기순이익"]),
                    ("자산총계", fin["자산총계"]),
                    ("부채총계", fin["부채총계"]),
                    ("자본총계", fin["자본총계"]),
                ]
                fin_headers = [f[0] for f in fin_items]
                fin_values = [fmt_money(f[1]) + "만원" if f[1] != 0 else "-" for f in fin_items]

                for c, h in enumerate(fin_headers, 1):
                    cell = ws0.cell(row=row, column=c, value=h)
                    cell.font = Font(bold=True, size=9, color="666666", name="Malgun Gothic")
                    cell.border = BD
                row += 1
                for c, v in enumerate(fin_values, 1):
                    cell = ws0.cell(row=row, column=c, value=v)
                    cell.font = DF
                    cell.border = BD

                # 부채비율 계산
                if fin["자본총계"] > 0:
                    debt_ratio = round(fin["부채총계"] / fin["자본총계"] * 100, 1)
                    ws0.cell(row=row, column=7, value=f"부채비율: {debt_ratio}%").font = Font(
                        bold=True, size=10, name="Malgun Gothic",
                        color="C00000" if debt_ratio > 200 else "006600"
                    )
                row += 1
            else:
                ws0.cell(row=row, column=1, value="재무정보: 법인등록번호 입력 시 자동 조회됩니다").font = Font(size=10, color="999999", italic=True, name="Malgun Gothic")
                row += 1

            # 매칭 발주
            ws0.cell(row=row, column=1, value=f"매칭 가능 재량계약: {len(mc)}건").font = Font(bold=True, size=10, name="Malgun Gothic")
            row += 1

            if mc:
                mh = ["사업명", "발주기관", "예산(만원)", "업종", "지역", "발주시기"]
                for c, h in enumerate(mh, 1):
                    cell = ws0.cell(row=row, column=c, value=h)
                    cell.fill, cell.font, cell.border = PatternFill("solid", fgColor="D6E4F0"), Font(bold=True, size=9, name="Malgun Gothic"), BD
                row += 1
                for m in mc[:15]:  # 최대 15건
                    for c, h in enumerate(mh, 1):
                        cell = ws0.cell(row=row, column=c, value=m.get(h, ""))
                        cell.font, cell.border = DF, BD
                    ws0.cell(row=row, column=3).number_format = "#,##0"
                    row += 1
                if len(mc) > 15:
                    ws0.cell(row=row, column=1, value=f"  ... 외 {len(mc)-15}건 (전체목록은 '고객매칭' 시트 참조)").font = Font(size=9, color="888888", italic=True, name="Malgun Gothic")
                    row += 1

            row += 2  # 고객사 간 간격

    else:
        ws0.cell(row=4, column=1, value="config.txt에 CLIENTS를 입력하면 고객사별 수주역량 진단이 여기에 표시됩니다.").font = Font(size=11, color="888888", italic=True, name="Malgun Gothic")

    aw(ws0)

    # ═══════════════════════════════════════════
    # 시트2: 재량계약 타겟
    # ═══════════════════════════════════════════
    ws1 = wb.create_sheet("재량계약 타겟")
    ws1.sheet_properties.tabColor = "70AD47"

    disc_sorted = sorted(disc_data, key=lambda x: x["예산액"], reverse=True)
    c1 = ["사업명", "발주기관", "예산(만원)", "업종", "재량여부", "지역", "발주시기", "담당자", "연락처"]
    for c, h in enumerate(c1, 1):
        ws1.cell(row=1, column=c, value=h)
    hdr(ws1, 1, len(c1))

    for r, d in enumerate(disc_sorted, 2):
        for c, h in enumerate(c1, 1):
            cell = ws1.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = DF, BD
            if r % 2 == 0:
                cell.fill = AF
        ws1.cell(row=r, column=3).number_format = "#,##0"
        if d["재량여부"] == "소액수의":
            ws1.cell(row=r, column=5).fill = GF

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(c1))}1"
    ws1.freeze_panes = "A2"
    aw(ws1)

    # ═══════════════════════════════════════════
    # 시트3: 전체 발주계획
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("전체 발주계획")
    ws2.sheet_properties.tabColor = "4472C4"

    c2 = ["번호", "사업명", "발주기관", "수요기관", "예산(만원)", "예산(억)",
          "발주시기", "계약방법", "업종", "재량여부", "지역", "담당자", "연락처"]
    for c, h in enumerate(c2, 1):
        ws2.cell(row=1, column=c, value=h)
    hdr(ws2, 1, len(c2))

    for r, d in enumerate(all_data, 2):
        for c, h in enumerate(c2, 1):
            cell = ws2.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = DF, BD
            if r % 2 == 0:
                cell.fill = AF
        ws2.cell(row=r, column=5).number_format = "#,##0"
        ws2.cell(row=r, column=6).number_format = "#,##0.00"
        if ws2.cell(row=r, column=10).value in ("소액수의", "재량계약"):
            ws2.cell(row=r, column=10).fill = GF

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(c2))}1"
    ws2.freeze_panes = "A2"
    aw(ws2)

    # ═══════════════════════════════════════════
    # 시트4: 통계
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("통계")
    ws3.sheet_properties.tabColor = "7030A0"

    ws3["A1"] = f"발주계획 분석 ({config.get('YEAR', '')})"
    ws3["A1"].font = TF
    ws3.merge_cells("A1:E1")
    ws3["A2"] = f"전체 {len(all_data)}건 | 재량계약 {len(disc_data)}건 | {datetime.now():%Y-%m-%d %H:%M}"
    ws3["A2"].font = Font(size=10, color="666666", name="Malgun Gothic")

    # 업종별
    cat_st = {}
    for d in all_data:
        ct = d["업종"]
        if ct not in cat_st:
            cat_st[ct] = {"n": 0, "amt": 0, "disc": 0}
        cat_st[ct]["n"] += 1
        cat_st[ct]["amt"] += d["예산액"]
        if d["재량여부"] in ("소액수의", "재량계약"):
            cat_st[ct]["disc"] += 1

    row = 4
    sc = ["업종", "건수", "금액(억)", "재량건수", "재량비율(%)"]
    for c, h in enumerate(sc, 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, len(sc))

    for ct, st in sorted(cat_st.items(), key=lambda x: x[1]["n"], reverse=True):
        row += 1
        ws3.cell(row=row, column=1, value=ct).font = DF
        ws3.cell(row=row, column=2, value=st["n"]).font = DF
        ws3.cell(row=row, column=3, value=round(st["amt"] / 1_0000_0000, 1)).font = DF
        ws3.cell(row=row, column=4, value=st["disc"]).font = DF
        pct = round(st["disc"] / st["n"] * 100, 1) if st["n"] > 0 else 0
        ws3.cell(row=row, column=5, value=pct).font = DF
        for c in range(1, 6):
            ws3.cell(row=row, column=c).border = BD

    # 지역별
    row += 2
    ws3.cell(row=row, column=1, value="[지역별]").font = SF
    row += 1
    for c, h in enumerate(["지역", "건수", "재량건수"], 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, 3)

    rgn = {}
    for d in all_data:
        r2 = d["지역"] or "미분류"
        if r2 not in rgn:
            rgn[r2] = {"n": 0, "d": 0}
        rgn[r2]["n"] += 1
        if d["재량여부"] in ("소액수의", "재량계약"):
            rgn[r2]["d"] += 1
    for rg, st in sorted(rgn.items(), key=lambda x: x[1]["n"], reverse=True)[:15]:
        row += 1
        ws3.cell(row=row, column=1, value=rg).font = DF
        ws3.cell(row=row, column=2, value=st["n"]).font = DF
        ws3.cell(row=row, column=3, value=st["d"]).font = DF

    aw(ws3)

    # ═══════════════════════════════════════════
    # 시트5: 고객매칭 (전체)
    # ═══════════════════════════════════════════
    if matched:
        ws4 = wb.create_sheet("고객매칭")
        ws4.sheet_properties.tabColor = "ED7D31"
        c4 = ["매칭고객사", "고객업종", "사업명", "발주기관", "예산(만원)", "재량여부", "지역", "담당자", "연락처"]
        for c, h in enumerate(c4, 1):
            ws4.cell(row=1, column=c, value=h)
        hdr(ws4, 1, len(c4))
        for r, d in enumerate(matched, 2):
            for c, h in enumerate(c4, 1):
                cell = ws4.cell(row=r, column=c, value=d.get(h, ""))
                cell.font, cell.border = DF, BD
            ws4.cell(row=r, column=5).number_format = "#,##0"
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(c4))}1"
        ws4.freeze_panes = "A2"
        aw(ws4)

    # ═══════════════════════════════════════════
    # 시트6: 고객사 재무현황
    # ═══════════════════════════════════════════
    if financials:
        ws5 = wb.create_sheet("고객 재무현황")
        ws5.sheet_properties.tabColor = "C00000"
        c5 = ["고객사", "업종", "기준연도", "매출액(만원)", "영업이익(만원)", "당기순이익(만원)",
              "자산총계(만원)", "부채총계(만원)", "자본총계(만원)", "부채비율(%)"]
        for c, h in enumerate(c5, 1):
            ws5.cell(row=1, column=c, value=h)
        hdr(ws5, 1, len(c5))

        r = 2
        for cn, fin in financials.items():
            cl = next((c for c in clients if c["name"] == cn), {})
            ws5.cell(row=r, column=1, value=cn).font = DF
            ws5.cell(row=r, column=2, value=cl.get("category", "")).font = DF
            ws5.cell(row=r, column=3, value=fin["기준연도"]).font = DF
            ws5.cell(row=r, column=4, value=round(fin["매출액"] / 10000) if fin["매출액"] else 0).font = DF
            ws5.cell(row=r, column=5, value=round(fin["영업이익"] / 10000) if fin["영업이익"] else 0).font = DF
            ws5.cell(row=r, column=6, value=round(fin["당기순이익"] / 10000) if fin["당기순이익"] else 0).font = DF
            ws5.cell(row=r, column=7, value=round(fin["자산총계"] / 10000) if fin["자산총계"] else 0).font = DF
            ws5.cell(row=r, column=8, value=round(fin["부채총계"] / 10000) if fin["부채총계"] else 0).font = DF
            ws5.cell(row=r, column=9, value=round(fin["자본총계"] / 10000) if fin["자본총계"] else 0).font = DF
            dr = round(fin["부채총계"] / fin["자본총계"] * 100, 1) if fin["자본총계"] > 0 else 0
            cell = ws5.cell(row=r, column=10, value=dr)
            cell.font = Font(bold=True, size=10, name="Malgun Gothic", color="C00000" if dr > 200 else "006600")
            for c in range(1, 11):
                ws5.cell(row=r, column=c).border = BD
                if c >= 4:
                    ws5.cell(row=r, column=c).number_format = "#,##0"
            r += 1
        aw(ws5)

    # 저장
    fname = f"G2B_Report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    fpath = OUTPUT_DIR / fname
    wb.save(str(fpath))
    return fpath


# =============================================================
#  메인
# =============================================================

def main(non_interactive=False):
    interactive = not non_interactive

    print()
    print("=" * 56)
    print("  건설업 수주역량 자동진단 시스템 v3.0")
    print("  행정사사무소 하랑 · 서울건설정보")
    print("=" * 56)
    print()

    # ── 설정 ──
    config = load_config()
    if config is None:
        print("  config.txt를 생성했습니다.")
        if not interactive:
            print("  [X] 비대화 모드에서는 최초 설정을 진행할 수 없습니다.")
            print(f"      config.txt를 먼저 작성하세요: {CONFIG_FILE}")
            return 1
        print("  인코딩/디코딩 키를 입력하세요.")
        print()
        if sys.platform == "win32":
            os.system(f'notepad "{CONFIG_FILE}"')
        wait_for_enter("  키 입력 후 Enter...", interactive=True)
        config = load_config()

    enc_key = config.get("ENCODING_KEY", "")
    dec_key = config.get("DECODING_KEY", "")

    if (not enc_key or "여기에" in enc_key) and (not dec_key or "여기에" in dec_key):
        print("  [X] API 키가 없습니다!")
        print(f"      config.txt: {CONFIG_FILE}")
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    log.info("  설정 OK")

    clients = parse_clients(config)
    if clients:
        log.info(f"  고객사 {len(clients)}곳: {', '.join(c['name'] for c in clients)}")

    # ── 1. 나라장터 발주계획 수집 ──
    log.info("")
    log.info("  [1/4] 나라장터 발주계획 수집")
    log.info("")

    raw = []
    for ep in ["const", "serv"]:
        items = g2b_fetch_all(enc_key, dec_key, ep)
        raw.extend(items)
        time.sleep(1)

    if not raw:
        print()
        print("  [X] 나라장터 데이터 수집 실패")
        print("  - 공공데이터포털 > 마이페이지 > 인증키 확인")
        print("  - '발주계획현황서비스' 활용신청 여부 확인")
        print("  - 키 포함 호출에서 HTTP 404 반복 시 활용신청 미승인/키 오류 가능성이 큽니다")
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    # ── 2. 데이터 처리 ──
    log.info("")
    log.info("  [2/4] 데이터 처리")

    all_data = process_g2b(raw, config)
    disc_data = [d for d in all_data if d["재량여부"] in ("소액수의", "재량계약")]

    cat_n = {}
    for d in all_data:
        cat_n[d["업종"]] = cat_n.get(d["업종"], 0) + 1

    log.info(f"  전체: {len(all_data)}건 | 재량계약: {len(disc_data)}건")
    for ct, n in sorted(cat_n.items(), key=lambda x: x[1], reverse=True):
        log.info(f"    {ct}: {n}건")

    # ── 3. 고객사 재무정보 + 매칭 ──
    log.info("")
    log.info("  [3/4] 고객사 분석")

    financials = {}
    matched = []

    if clients:
        matched = match_clients(disc_data, clients)
        log.info(f"  발주 매칭: {len(matched)}건")

        # 법인등록번호가 있는 고객사만 재무조회
        crno_clients = [c for c in clients if c.get("crno")]
        if crno_clients:
            log.info(f"  재무정보 조회: {len(crno_clients)}사")
            for cl in crno_clients:
                log.info(f"    {cl['name']}...")
                fin = get_client_financials(enc_key, dec_key, cl)
                if fin:
                    financials[cl["name"]] = fin
                    log.info(f"    -> {fin['기준연도']}년 매출 {fmt_eok_simple(fin['매출액'])}억")
                else:
                    log.info(f"    -> 조회 실패 (법인번호 확인 필요)")
                time.sleep(0.5)
        else:
            log.info("  법인등록번호가 있는 고객사 없음 - 재무조회 건너뜀")
    else:
        log.info("  등록된 고객사 없음")

    # ── 4. 리포트 생성 ──
    log.info("")
    log.info("  [4/4] 리포트 생성")

    if not ensure_openpyxl():
        wait_for_enter("\n  Enter로 종료...", interactive=interactive)
        return 1

    fpath = create_report(all_data, disc_data, matched, clients, financials, config)

    # ── 완료 ──
    print()
    print("=" * 56)
    print(f"  [OK] 완료!")
    print(f"  파일: {fpath}")
    print(f"  전체 {len(all_data)}건 | 재량계약 {len(disc_data)}건")
    if matched:
        print(f"  고객매칭 {len(matched)}건")
    if financials:
        print(f"  재무조회 {len(financials)}사")
    print("=" * 56)
    print()

    if interactive and sys.platform == "win32":
        os.startfile(str(OUTPUT_DIR))

    wait_for_enter("  Enter로 종료...", interactive=interactive)
    return 0


def fmt_eok_simple(v):
    if v == 0:
        return "0"
    return f"{v / 1_0000_0000:,.1f}"


def parse_args():
    parser = argparse.ArgumentParser(description="G2B 발주계획 자동 수집 및 리포트 생성")
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="입력 대기/메모장 열기/결과 폴더 자동 열기를 비활성화합니다.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        raise SystemExit(main(non_interactive=args.non_interactive))
    except KeyboardInterrupt:
        print("\n  중단")
        wait_for_enter("\n  Enter로 종료...", interactive=not args.non_interactive)
        raise SystemExit(130)
    except Exception as e:
        log.error(f"  오류: {e}")
        import traceback
        traceback.print_exc()
        wait_for_enter("\n  Enter로 종료...", interactive=not args.non_interactive)
        raise SystemExit(1)
