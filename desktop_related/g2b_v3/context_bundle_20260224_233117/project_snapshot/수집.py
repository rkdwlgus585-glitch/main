#!/usr/bin/env python3
"""
=============================================================
  나라장터 발주계획 자동수집 시스템 v2.0
  행정사사무소 하랑 · 서울건설정보
=============================================================
  
  ★ 설정 방법: config.txt 파일에 API키만 붙여넣으세요
  ★ 실행 방법: '실행.bat' 더블클릭
  
=============================================================
"""

import os, sys, json, time, logging
from datetime import datetime, timedelta
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.parse import urlencode
from urllib.error import URLError

# ── 경로 설정 (실행파일 기준) ──
BASE_DIR = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "config.txt"
OUTPUT_DIR = BASE_DIR / "결과"
LOG_DIR = BASE_DIR / "로그"

# ── 로깅 ──
LOG_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_DIR / f"실행로그_{datetime.now():%Y%m%d}.log", encoding="utf-8"),
    ]
)
log = logging.getLogger()


# =============================================================
#  설정 로드
# =============================================================

def load_config():
    """config.txt에서 설정 로드"""
    config = {
        "API_KEY": "",
        "조회년도": str(datetime.now().year),
        "타겟업종": "전기,정보통신,소방,건축,토목,기계설비,조경",
        "재량계약기준액": "20000",  # 만원 단위 (2억)
        "소액수의기준액": "5000",   # 만원 단위 (5천만원)
    }
    
    if not CONFIG_FILE.exists():
        # 최초 실행 - config.txt 자동 생성
        lines = [
            "# ============================================",
            "#  나라장터 발주계획 자동수집 설정파일",
            "#  행정사사무소 하랑 · 서울건설정보",
            "# ============================================",
            "#",
            "# ★ 아래 API_KEY= 뒤에 공공데이터포털 API키를 붙여넣으세요",
            "#   (https://www.data.go.kr/data/15129462/openapi.do)",
            "#",
            "",
            "API_KEY=여기에_API키_붙여넣기",
            "",
            "# ── 아래는 수정하지 않아도 됩니다 ──",
            f"조회년도={config['조회년도']}",
            f"타겟업종={config['타겟업종']}",
            f"재량계약기준액={config['재량계약기준액']}",
            f"소액수의기준액={config['소액수의기준액']}",
            "",
            "# ── 인큐베이팅 고객사 (회사명:업종 형태로 추가) ──",
            "# 예시: 고객사=하늘건설:일반건설,동방전기:전기공사,서울ICT:정보통신",
            "고객사=",
        ]
        CONFIG_FILE.write_text("\n".join(lines), encoding="utf-8")
        return None  # 최초 실행 표시
    
    for line in CONFIG_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            key, val = line.split("=", 1)
            config[key.strip()] = val.strip()
    
    return config


def parse_clients(config):
    """고객사 문자열 파싱"""
    raw = config.get("고객사", "")
    if not raw:
        return {}
    
    clients = {}
    for pair in raw.split(","):
        pair = pair.strip()
        if ":" in pair:
            name, cat = pair.split(":", 1)
            clients[name.strip()] = cat.strip()
    return clients


# =============================================================
#  API 호출 (requests 없이 순수 urllib)
# =============================================================

BASE_URL = "http://apis.data.go.kr/1230000/ao/OrderPlanSttusService"

ENDPOINTS = {
    "공사": "getOrderPlanSttusListCnstwk",
    "용역": "getOrderPlanSttusListServce",
    "물품": "getOrderPlanSttusListThng",
    "공사검색": "getOrderPlanSttusSearchListCnstwk",
    "용역검색": "getOrderPlanSttusSearchListServce",
}


def api_call(api_key, endpoint_key, page=1, rows=999):
    """API 1회 호출"""
    endpoint = ENDPOINTS[endpoint_key]
    params = urlencode({
        "serviceKey": api_key,
        "pageNo": str(page),
        "numOfRows": str(rows),
        "type": "json",
    })
    url = f"{BASE_URL}/{endpoint}?{params}"
    
    try:
        req = Request(url, headers={"Accept": "application/json"})
        with urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        
        if "response" in data:
            header = data["response"].get("header", {})
            if header.get("resultCode") != "00":
                log.error(f"  API 오류: {header.get('resultMsg', '알수없음')}")
                return [], 0
            body = data["response"].get("body", {})
            items = body.get("items", [])
            if isinstance(items, dict):
                items = items.get("item", [])
            if isinstance(items, dict):
                items = [items]
            total = int(body.get("totalCount", 0))
            return items, total
        return [], 0
    except URLError as e:
        log.error(f"  네트워크 오류: {e}")
        return [], 0
    except Exception as e:
        log.error(f"  API 호출 실패: {e}")
        return [], 0


def fetch_all(api_key, biz_type, max_pages=30):
    """전체 페이지 자동 수집"""
    all_items = []
    page = 1
    
    while page <= max_pages:
        log.info(f"  [{biz_type}] {page}페이지 수집중...")
        items, total = api_call(api_key, biz_type, page=page)
        
        if not items:
            break
        all_items.extend(items)
        
        if len(all_items) >= total:
            break
        page += 1
        time.sleep(0.3)
    
    log.info(f"  [{biz_type}] {len(all_items)}건 수집 완료 (전체 {total if 'total' in dir() else '?'}건)")
    return all_items


# =============================================================
#  데이터 처리
# =============================================================

CATEGORY_RULES = {
    "전기공사": ["전기", "전력", "배전", "수변전", "조명", "전기설비", "전기공사"],
    "정보통신공사": ["정보통신", "통신", "네트워크", "CCTV", "방송", "통합배선", "ICT", "cctv"],
    "소방공사": ["소방", "방재", "소화", "경보", "스프링클러", "소방설비"],
    "기계설비공사": ["기계설비", "냉난방", "공조", "배관", "급배수", "기계"],
    "조경공사": ["조경", "녹화", "수목", "식재"],
    "건축공사": ["건축", "신축", "증축", "리모델링", "개보수", "인테리어"],
    "토목공사": ["토목", "도로", "상하수도", "하천", "교량", "포장"],
}


def classify(name):
    """사업명으로 업종 자동 분류"""
    if not name:
        return "기타"
    upper = name.upper()
    for cat, keywords in CATEGORY_RULES.items():
        for kw in keywords:
            if kw.upper() in upper:
                return cat
    return "기타"


def judge_contract_type(amount, threshold_disc, threshold_small):
    """재량계약 여부 판단"""
    if amount <= 0:
        return "금액미정"
    if amount <= threshold_small:
        return "소액수의"
    if amount <= threshold_disc:
        return "재량계약"
    return "일반경쟁"


def process(raw_items, config):
    """원시 데이터 → 정제"""
    disc_limit = int(config.get("재량계약기준액", "20000")) * 10000
    small_limit = int(config.get("소액수의기준액", "5000")) * 10000
    
    results = []
    for item in raw_items:
        try:
            amt = int(item.get("asignBdgtAmt", 0) or 0)
        except (ValueError, TypeError):
            amt = 0
        
        name = item.get("bidNtceNm", "") or item.get("prdctClsfcNoNm", "") or ""
        
        results.append({
            "번호": item.get("orderPlanUntyNo", ""),
            "사업명": name,
            "발주기관": item.get("orderInsttNm", ""),
            "수요기관": item.get("dminsttNm", ""),
            "예산액": amt,
            "예산(만원)": round(amt / 10000) if amt > 0 else 0,
            "예산(억)": round(amt / 1_0000_0000, 2) if amt > 0 else 0,
            "발주시기": item.get("orderPlanRegDt", ""),
            "계약방법": item.get("cntrctMthdNm", ""),
            "업종": classify(name),
            "재량여부": judge_contract_type(amt, disc_limit, small_limit),
            "지역": item.get("dminsttRgnNm", "") or item.get("orderInsttRgnNm", ""),
            "담당자": item.get("orderInsttOfclNm", ""),
            "연락처": item.get("orderInsttOfclTelNo", ""),
        })
    
    return results


# =============================================================
#  엑셀 생성
# =============================================================

def create_excel(all_data, config, clients_map):
    """엑셀 리포트 생성 (openpyxl)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl이 설치되지 않았습니다. 자동 설치합니다...")
        os.system(f'"{sys.executable}" -m pip install openpyxl -q')
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # 스타일
    H_FILL = PatternFill("solid", fgColor="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
    D_FONT = Font(size=10, name="맑은 고딕")
    T_FONT = Font(bold=True, size=14, name="맑은 고딕", color="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")
    WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    
    def hdr(ws, row, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill, cell.font, cell.border = H_FILL, H_FONT, BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def auto_w(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 3, 10), 40)
    
    # ── 재량계약 타겟 (메인 시트) ──
    ws1 = wb.active
    ws1.title = "재량계약 타겟"
    ws1.sheet_properties.tabColor = "70AD47"
    
    disc_data = [d for d in all_data if d["재량여부"] in ("소액수의", "재량계약")]
    disc_data.sort(key=lambda x: x["예산액"], reverse=True)
    
    cols1 = ["사업명", "발주기관", "예산(만원)", "업종", "재량여부", "지역", "발주시기", "담당자", "연락처"]
    for c, h in enumerate(cols1, 1):
        ws1.cell(row=1, column=c, value=h)
    hdr(ws1, 1, len(cols1))
    
    for r, d in enumerate(disc_data, 2):
        for c, h in enumerate(cols1, 1):
            cell = ws1.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = D_FONT, BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
        # 예산 숫자 포맷
        ws1.cell(row=r, column=3).number_format = "#,##0"
        # 소액수의 강조
        if d["재량여부"] == "소액수의":
            ws1.cell(row=r, column=5).fill = OK_FILL
    
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols1))}1"
    ws1.freeze_panes = "A2"
    auto_w(ws1)
    
    # ── 전체 발주계획 ──
    ws2 = wb.create_sheet("전체 발주계획")
    ws2.sheet_properties.tabColor = "4472C4"
    
    cols2 = ["번호", "사업명", "발주기관", "수요기관", "예산(만원)", "예산(억)", 
             "발주시기", "계약방법", "업종", "재량여부", "지역", "담당자", "연락처"]
    for c, h in enumerate(cols2, 1):
        ws2.cell(row=1, column=c, value=h)
    hdr(ws2, 1, len(cols2))
    
    for r, d in enumerate(all_data, 2):
        for c, h in enumerate(cols2, 1):
            cell = ws2.cell(row=r, column=c, value=d.get(h, ""))
            cell.font, cell.border = D_FONT, BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
        ws2.cell(row=r, column=5).number_format = "#,##0"
        ws2.cell(row=r, column=6).number_format = "#,##0.00"
        
        rc = ws2.cell(row=r, column=10)
        if rc.value in ("소액수의", "재량계약"):
            rc.fill = OK_FILL
    
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
    ws2.freeze_panes = "A2"
    auto_w(ws2)
    
    # ── 업종별 통계 ──
    ws3 = wb.create_sheet("업종별 통계")
    ws3.sheet_properties.tabColor = "7030A0"
    
    ws3["A1"] = f"나라장터 발주계획 분석 ({config.get('조회년도', '')}년)"
    ws3["A1"].font = T_FONT
    ws3.merge_cells("A1:E1")
    ws3["A2"] = f"생성: {datetime.now():%Y-%m-%d %H:%M} | 전체 {len(all_data)}건 | 재량계약 {len(disc_data)}건"
    ws3["A2"].font = Font(size=10, color="666666", name="맑은 고딕")
    
    # 업종별 집계
    cat_stats = {}
    for d in all_data:
        cat = d["업종"]
        if cat not in cat_stats:
            cat_stats[cat] = {"건수": 0, "금액합계": 0, "재량건수": 0}
        cat_stats[cat]["건수"] += 1
        cat_stats[cat]["금액합계"] += d["예산액"]
        if d["재량여부"] in ("소액수의", "재량계약"):
            cat_stats[cat]["재량건수"] += 1
    
    row = 4
    stat_cols = ["업종", "전체건수", "금액합계(억)", "재량계약건수", "재량비율(%)"]
    for c, h in enumerate(stat_cols, 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, len(stat_cols))
    
    for cat, st in sorted(cat_stats.items(), key=lambda x: x[1]["건수"], reverse=True):
        row += 1
        ws3.cell(row=row, column=1, value=cat).font = D_FONT
        ws3.cell(row=row, column=2, value=st["건수"]).font = D_FONT
        ws3.cell(row=row, column=3, value=round(st["금액합계"] / 1_0000_0000, 1)).font = D_FONT
        ws3.cell(row=row, column=4, value=st["재량건수"]).font = D_FONT
        pct = round(st["재량건수"] / st["건수"] * 100, 1) if st["건수"] > 0 else 0
        ws3.cell(row=row, column=5, value=pct).font = D_FONT
        for c in range(1, 6):
            ws3.cell(row=row, column=c).border = BORDER
    
    # 지역별
    row += 3
    ws3.cell(row=row, column=1, value="■ 지역별 현황").font = Font(bold=True, size=12, name="맑은 고딕")
    row += 1
    for c, h in enumerate(["지역", "건수", "재량계약건수"], 1):
        ws3.cell(row=row, column=c, value=h)
    hdr(ws3, row, 3)
    
    rgn_stats = {}
    for d in all_data:
        rgn = d["지역"] or "미분류"
        if rgn not in rgn_stats:
            rgn_stats[rgn] = {"건수": 0, "재량": 0}
        rgn_stats[rgn]["건수"] += 1
        if d["재량여부"] in ("소액수의", "재량계약"):
            rgn_stats[rgn]["재량"] += 1
    
    for rgn, st in sorted(rgn_stats.items(), key=lambda x: x[1]["건수"], reverse=True)[:15]:
        row += 1
        ws3.cell(row=row, column=1, value=rgn).font = D_FONT
        ws3.cell(row=row, column=2, value=st["건수"]).font = D_FONT
        ws3.cell(row=row, column=3, value=st["재량"]).font = D_FONT
    
    auto_w(ws3)
    
    # ── 고객사 매칭 (고객사 있을 때만) ──
    if clients_map:
        ws4 = wb.create_sheet("고객사 매칭")
        ws4.sheet_properties.tabColor = "ED7D31"
        
        cols4 = ["매칭고객사", "고객업종", "사업명", "발주기관", "예산(만원)", "재량여부", "지역", "담당자", "연락처"]
        for c, h in enumerate(cols4, 1):
            ws4.cell(row=1, column=c, value=h)
        hdr(ws4, 1, len(cols4))
        
        match_row = 2
        for client_name, client_cat in clients_map.items():
            for d in disc_data:
                # 업종 매칭 로직
                item_cat = d["업종"]
                matched = False
                if "일반건설" in client_cat and ("건축" in item_cat or "토목" in item_cat):
                    matched = True
                elif "전기" in client_cat and "전기" in item_cat:
                    matched = True
                elif "정보통신" in client_cat and "정보통신" in item_cat:
                    matched = True
                elif "소방" in client_cat and "소방" in item_cat:
                    matched = True
                elif "기계" in client_cat and "기계" in item_cat:
                    matched = True
                elif "조경" in client_cat and "조경" in item_cat:
                    matched = True
                
                if matched:
                    vals = [client_name, client_cat, d["사업명"], d["발주기관"], 
                            d["예산(만원)"], d["재량여부"], d["지역"], d["담당자"], d["연락처"]]
                    for c, v in enumerate(vals, 1):
                        cell = ws4.cell(row=match_row, column=c, value=v)
                        cell.font, cell.border = D_FONT, BORDER
                    ws4.cell(row=match_row, column=5).number_format = "#,##0"
                    match_row += 1
        
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"
        ws4.freeze_panes = "A2"
        auto_w(ws4)
        log.info(f"  고객사 매칭: {match_row - 2}건")
    
    # 저장
    today = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"발주계획_{today}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    return filepath


# =============================================================
#  메인 실행
# =============================================================

def main():
    print()
    print("=" * 56)
    print("  🏗️  나라장터 발주계획 자동수집 시스템")
    print("  행정사사무소 하랑 · 서울건설정보")
    print("=" * 56)
    print()
    
    # 1. 설정 로드
    config = load_config()
    
    if config is None:
        print("📌 최초 실행 - config.txt 파일이 생성되었습니다!")
        print()
        print("   ➜ config.txt를 메모장으로 열어서")
        print("     API_KEY= 뒤에 공공데이터포털 API키를 붙여넣으세요")
        print()
        print(f"   파일 위치: {CONFIG_FILE}")
        print()
        
        # 바로 메모장으로 열기
        if sys.platform == "win32":
            os.system(f'notepad "{CONFIG_FILE}"')
        
        input("API키 입력 후 Enter를 누르면 계속합니다...")
        config = load_config()
    
    api_key = config.get("API_KEY", "")
    if not api_key or api_key == "여기에_API키_붙여넣기":
        print("❌ API 키가 설정되지 않았습니다!")
        print(f"   config.txt를 열어 API_KEY를 입력하세요: {CONFIG_FILE}")
        input("\nEnter를 누르면 종료합니다...")
        return
    
    log.info(f"✅ 설정 로드 완료 (조회년도: {config.get('조회년도')})")
    
    clients_map = parse_clients(config)
    if clients_map:
        log.info(f"  인큐베이팅 고객: {', '.join(clients_map.keys())}")
    
    # 2. 데이터 수집
    log.info("")
    log.info("📡 나라장터 API 데이터 수집 시작...")
    log.info("")
    
    raw_items = []
    for biz in ["공사", "용역"]:
        items = fetch_all(api_key, biz)
        raw_items.extend(items)
        time.sleep(1)
    
    if not raw_items:
        print()
        print("❌ 수집된 데이터가 없습니다.")
        print("   - API 키가 올바른지 확인하세요")
        print("   - 네트워크 연결을 확인하세요")
        print("   - 공공데이터포털에서 '발주계획현황서비스' 활용신청 여부를 확인하세요")
        input("\nEnter를 누르면 종료합니다...")
        return
    
    # 3. 데이터 처리
    log.info("")
    log.info("🔧 데이터 처리 및 분류...")
    all_data = process(raw_items, config)
    
    disc_count = len([d for d in all_data if d["재량여부"] in ("소액수의", "재량계약")])
    
    # 업종별 카운트
    cat_count = {}
    for d in all_data:
        cat_count[d["업종"]] = cat_count.get(d["업종"], 0) + 1
    
    log.info(f"  전체: {len(all_data)}건")
    log.info(f"  재량계약 가능: {disc_count}건 ({round(disc_count/len(all_data)*100, 1) if all_data else 0}%)")
    for cat, cnt in sorted(cat_count.items(), key=lambda x: x[1], reverse=True):
        log.info(f"    {cat}: {cnt}건")
    
    # 4. 엑셀 생성
    log.info("")
    log.info("📊 엑셀 리포트 생성중...")
    filepath = create_excel(all_data, config, clients_map)
    
    # 5. 완료
    print()
    print("=" * 56)
    print(f"  ✅ 완료!")
    print(f"  📁 파일: {filepath}")
    print(f"  📊 전체 {len(all_data)}건 | 재량계약 {disc_count}건")
    if clients_map:
        print(f"  🤝 고객사 {len(clients_map)}개 매칭 포함")
    print("=" * 56)
    print()
    
    # 결과 폴더 열기
    if sys.platform == "win32":
        os.startfile(str(OUTPUT_DIR))
    
    input("Enter를 누르면 종료합니다...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n중단됨")
    except Exception as e:
        log.error(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
        input("\nEnter를 누르면 종료합니다...")
